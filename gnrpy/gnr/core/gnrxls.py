from builtins import str
#from builtins import object
import xlwt
import six
try:
    import openpyxl
except:
    import sys
    print("\n**ERROR Missing openpyxl: can't export in xlsx\n", file=sys.stderr)
import os
from gnr.core.gnrstring import toText
from gnr.lib.services.storage import StorageNode

class BaseXls(object):
    
    def save(self,filepath=None,autosize=None):
        self.filepath = filepath
        if autosize is None:
            autosize = True
        self.workbookSave(autosize=autosize)

    def getSheet(self,name=None):
        name = name or self.sheet_base_name
        if not self.sheets:
            self.createSheet(self.sheet_base_name,headers=self.sheet_base_headers,columns=self.sheet_base_columns,
                coltypes=self.sheet_base_coltypes, groups=self.sheet_base_groups)
        return self.sheets[name]
    
    @property
    def sheet(self):
        return self.getSheet()['sheet']
        
    @property
    def headers(self):
        return self.getSheet()['headers']

    @property
    def columns(self):
        return self.getSheet()['columns']

    @property
    def colsizes(self):
        return  self.getSheet()['colsizes']

    @property
    def coltypes(self):
        return self.getSheet()['coltypes']

    @property
    def current_row(self):
        return self.getSheet()['current_row']

    @property
    def groups(self):
        return  self.getSheet()['groups']


    @property
    def filepath(self):
        return self._filepath

    @filepath.setter
    def filepath(self, filepath):
        if filepath is None:
            self._filepath = None
            self.filenode = None
            return
        self.filenode = None
        if isinstance(filepath, StorageNode):
            self.filenode = filepath
            filepath = self.filenode.path
        filepath = f'{os.path.splitext(filepath)[0]}.{self.extension}'
        if self.filenode:
            self.filenode.path = filepath
        self._filepath = filepath

    def composeAll(self,data=None,**kwargs):
        for export_data in data:
            self.compose(export_data)
    
    def compose(self,data):
        sheet_name = data['name'].replace('/','')
        self.createSheet(sheet_name,**data['struct'])
        self.writeHeaders(sheet_name=sheet_name)
        for row in data['rows']:
            self.writeRow(row,sheet_name=sheet_name)




class XlsWriter(BaseXls):
    """TODO"""

    extension = 'xls'
    content_type = 'application/xls'

    def __init__(self, columns=None, coltypes=None, headers=None, groups=None, filepath=None,sheet_base_name=None,
                 font='Times New Roman', format_float='#,##0.00', format_int='#,##0', locale=None):
       #self.headers = headers
       #self.columns = columns
        self.sheets = {}
        self.filepath = filepath
        if sheet_base_name is not False:
            self.sheet_base_headers = headers
            self.sheet_base_groups = groups
            self.sheet_base_coltypes = coltypes
            self.sheet_base_columns = columns
            self.sheet_base_name = sheet_base_name or os.path.basename(self.filepath)[:31] if self.filepath else '_base_'
        else:
            self.sheet_base_name = False
        #self.sheet = self.workbook.add_sheet(os.path.basename(self.filepath)[:31])
        self.locale = locale
        self.float_style = xlwt.XFStyle()
        self.float_style.num_format_str = format_float
        self.int_style = xlwt.XFStyle()
        self.int_style.num_format_str = format_int
        font0 = xlwt.Font()
        font0.name = font  # FIXED
        font0.bold = True
        
        self.hstyle = xlwt.XFStyle()
        self.hstyle.font = font0

    def __call__(self, data=None, sheet_name=None):
        self.writeHeaders(sheet_name=sheet_name)
        for item in data:
            row = self.rowGetter(item)
            self.writeRow(row,sheet_name=sheet_name)
        self.workbookSave()
    
    def createSheet(self,sheetname,headers=None,columns=None,coltypes=None,colsizes=None,
            groups=None):
        colsizes = colsizes or dict()
        self.sheets[sheetname] = {'sheet': self.workbook.add_sheet(sheetname),
                                   'headers':headers,'columns':columns,
                                    'colsizes':colsizes,'coltypes':coltypes,
                                    'groups':groups}
        self.sheets[sheetname]['sheet'].panes_frozen = True
        self.sheets[sheetname]['sheet'].horz_split_pos = 1 if not groups else 2
        
    

    def rowGetter(self, item):
        """TODO
        
        :param item: TODO"""
        return dict(item)
        
    def writeHeaders(self,sheet_name=None):
        """TODO"""
        sheet_name = sheet_name or self.sheet_base_name
        sheet_obj = self.getSheet(sheet_name)
        sheet = sheet_obj['sheet']
        headers = sheet_obj['headers']
        colsizes = sheet_obj['colsizes']
        groups = sheet_obj['groups']
        current_row = 0
        if groups:
            group_style = xlwt.easyxf('align: wrap on, vert centre, horiz center')
            current_col = 0
            for g in groups:
                name = g.get('name')
                start = g.get('start')
                end = g.get('end')
                sheet.write_merge(current_row, current_row,
                    start, end, name, group_style)
            current_row +=1
        for c, header in enumerate(headers):
            sheet.write(current_row, c, header, self.hstyle)
            colsizes[c] = max(colsizes.get(c, 0), self.fitwidth(header))
        sheet_obj['current_row'] = current_row

    def workbookSave(self,**kwargs):
        """TODO"""
        if self.filenode:
            with self.filenode.open(mode='wb') as outfile:
                self.workbook.save(outfile)
        else:
            self.workbook.save(self.filepath)
    
    def write(self,sheet_name=None,what=None):
        pass


    def writeRow(self, row, sheet_name=None):
        """TODO
        
        :param row: TODO"""
        sheet_name = sheet_name or self.sheet_base_name
        sheet_obj = self.getSheet(sheet_name)

        current_row = sheet_obj['current_row'] + 1
        sheet_obj['current_row'] = current_row
        sheet = sheet_obj['sheet']
        columns = sheet_obj['columns']
        coltypes = sheet_obj['coltypes']
        colsizes = sheet_obj['colsizes']

        for c, col in enumerate(columns):
            value = row.get(col)
            if isinstance(value, list):
                value = ','.join([str(x != None and x or '') for x in value])
            coltype = coltypes.get(col)
            if coltype in ('R', 'F', 'N'):
                sheet.write(current_row, c, value, self.float_style)
            elif coltype in ('L', 'I'):
                sheet.write(current_row, c, value, self.int_style)
            else:
                value = toText(value, self.locale)
                sheet.write(current_row, c, value)
            colsizes[c] = max(colsizes.get(c, 0), self.fitwidth(value))
            
    def fitwidth(self, data, bold=False):
        """Try to autofit Arial 10
        
        :param data: TODO
        :param bold: TODO"""
        charwidths = {
            '0': 262.637, '1': 262.637, '2': 262.637, '3': 262.637, '4': 262.637, '5': 262.637, '6': 262.637,
            '7': 262.637, '8': 262.637, '9': 262.637, 'a': 262.637, 'b': 262.637, 'c': 262.637, 'd': 262.637,
            'e': 262.637, 'f': 146.015, 'g': 262.637, 'h': 262.637, 'i': 117.096, 'j': 88.178, 'k': 233.244,
            'l': 88.178, 'm': 379.259, 'n': 262.637, 'o': 262.637, 'p': 262.637, 'q': 262.637, 'r': 175.407,
            's': 233.244, 't': 117.096, 'u': 262.637, 'v': 203.852, 'w': 321.422, 'x': 203.852, 'y': 262.637,
            'z': 233.244, 'A': 321.422, 'B': 321.422, 'C': 350.341, 'D': 350.341, 'E': 321.422, 'F': 291.556,
            'G': 350.341, 'H': 321.422, 'I': 146.015, 'J': 262.637, 'K': 321.422, 'L': 262.637, 'M': 379.259,
            'N': 321.422, 'O': 350.341, 'P': 321.422, 'Q': 350.341, 'R': 321.422, 'S': 321.422, 'T': 262.637,
            'U': 321.422, 'V': 321.422, 'W': 496.356, 'X': 321.422, 'Y': 321.422, 'Z': 262.637, ' ': 146.015,
            '!': 146.015, '"': 175.407, '#': 262.637, '$': 262.637, '%': 438.044, '&': 321.422, '\'': 88.178,
            '(': 175.407, ')': 175.407, '*': 203.852, '+': 291.556, ',': 146.015, '-': 175.407, '.': 146.015,
            '/': 146.015, ':': 146.015, ';': 146.015, '<': 291.556, '=': 291.556, '>': 291.556, '?': 262.637,
            '@': 496.356, '[': 146.015, '\\': 146.015, ']': 146.015, '^': 203.852, '_': 262.637, '`': 175.407,
            '{': 175.407, '|': 146.015, '}': 175.407, '~': 291.556}
        units = 220
        try:
            data = str(data)
        except UnicodeEncodeError:
            return max(units*len(data), 700)
        for char in data:
            if char in charwidths:
                units += charwidths[char]
            else:
                units += charwidths['0']
        if bold:
            units *= 1.1
        return max(units, 700) # Don't go smaller than a reported width of 2
            
class XlsReader(object):
    """docstring for XlsReader"""
    def __init__(self, arg):
        self.arg = arg  



class XlsxWriter(BaseXls):
    """TODO"""

    extension = 'xlsx'
    content_type = 'application/xlsx'


    def __init__(self, columns=None, coltypes=None, headers=None, groups=None, filepath=None,sheet_base_name=None,
                 font='Times New Roman', format_float='#,##0.00', format_int='#,##0', locale=None):
       #self.headers = headers
       #self.columns = columns
        self.sheets = {}
        self.filepath = filepath
        self.workbook = openpyxl.Workbook()  # self.workbook.active
        del self.workbook[self.workbook.sheetnames[0]]  # elimino il primo
        
        if sheet_base_name is not False:       
            self.sheet_base_name = sheet_base_name or os.path.basename(self.filepath)[:31] if self.filepath else '_base_'
            if sheet_base_name is not False:
                self.sheet_base_headers = headers
                self.sheet_base_groups = groups
                self.sheet_base_coltypes = coltypes
                self.sheet_base_columns = columns
        else:
            self.sheet_base_name = False

        #self.sheet = self.workbook.add_sheet(os.path.basename(self.filepath)[:31])
        self.locale = locale
        self.workbook.add_named_style(openpyxl.styles.NamedStyle('float',
                                font=openpyxl.styles.Font(name=font),
                                number_format=format_float,
                                alignment=openpyxl.styles.Alignment(vertical="top")
        ))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle('int',
                                font=openpyxl.styles.Font(name=font),
                                number_format=format_int,
                                alignment=openpyxl.styles.Alignment(vertical="top")
        ))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle('header',
                                font=openpyxl.styles.Font(name=font, bold=True),
                                alignment=openpyxl.styles.Alignment(horizontal="center", vertical="top")
        ))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle("default",
                                font=openpyxl.styles.Font(name=font),
                                alignment=openpyxl.styles.Alignment(vertical="top")
        ))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle("date",
                                font=openpyxl.styles.Font(name=font),
                                number_format="D MMM YYYY",
                                alignment=openpyxl.styles.Alignment(vertical="top"),
        ))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle("datetime",
                                font=openpyxl.styles.Font(name=font),
                                number_format="D MMM YYYY, H:MM:SS",
                                alignment=openpyxl.styles.Alignment(vertical="top"),
        ))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle("group",
                                font=openpyxl.styles.Font(name=font, bold=True),  # italic=True 
                                alignment=openpyxl.styles.Alignment(
                                    vertical="center", 
                                    wrap_text=True,                                        
                                    horizontal='center'
                                )
        ))


    def createSheet(self,sheetname,headers=None,columns=None,coltypes=None,colsizes=None,
            groups=None):
        colsizes = colsizes or dict()
        self.sheets[sheetname] = {'sheet': self.workbook.create_sheet(title=sheetname),
                                   'headers':headers,'columns':columns,
                                    'colsizes':colsizes,'coltypes':coltypes,
                                    'groups':groups}
        self.sheets[sheetname]['sheet'].panes_frozen = True
        self.sheets[sheetname]['sheet'].horz_split_pos = 1 if not groups else 2

    def __call__(self, data=None, sheet_name=None):
        self.writeHeaders(sheet_name=sheet_name)
        for item in data:
            row = self.rowGetter(item)
            self.writeRow(row,sheet_name=sheet_name)
        self.workbookSave()

    def rowGetter(self, item):
        """TODO

        :param item: TODO"""
        return dict(item)

    def writeHeaders(self,sheet_name=None):
        """TODO"""
        sheet_name = sheet_name or self.sheet_base_name
        sheet_obj = self.getSheet(sheet_name)
        sheet = sheet_obj['sheet']
        headers = sheet_obj['headers']
        colsizes = sheet_obj['colsizes']
        groups = sheet_obj['groups']
        current_row = 0
        if groups:
            current_col = 0
            for g in groups:
                name = g.get('name')
                start = g.get('start')
                end = g.get('end')
                self.writeCell(sheet, current_row, current_row+start,
                               name, style='group',
                               end_column=current_row+end)
            current_row +=1
        max_height = 0
        for c, header in enumerate(headers):
            self.writeCell(sheet, current_row, c, header, style="header")
            heigth, width = self.fitwidth(header)
            max_height = max(max_height, heigth)
            colsizes[c] = max(colsizes.get(c, 0), width)
        sheet_obj['current_row'] = current_row
        sheet.row_dimensions[current_row+1].height = max_height
    
    def workbookSave(self, autosize=True):
        """TODO"""

        if autosize==True:
            for sheetname in self.sheets:
                ws = self.workbook[sheetname]
                for col,size in self.sheets[sheetname]['colsizes'].items():
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col+1)].width = size

        if self.filenode:
            with self.filenode.open(mode='wb') as outfile:
                self.workbook.save(filename=outfile)
        else:
            self.workbook.save(filename=self.filepath)

    def writeCell(self, sheet, row, column, value, style=None, end_column=None):
        cell = sheet.cell(row=row+1, column=column+1, value=value)
        if style:
            cell.style = style
        if end_column:  # merging cells
            sheet.merge_cells(start_row=row+1, start_column=column+1,
                              end_row=row+1, end_column=end_column+1)
        return cell

    def writeRow(self, row, sheet_name=None):
        """TODO

        :param row: TODO"""
        sheet_name = sheet_name or self.sheet_base_name
        sheet_obj = self.getSheet(sheet_name)
        current_row = sheet_obj['current_row'] + 1
        sheet_obj['current_row'] = current_row
        sheet = sheet_obj['sheet']
        columns = sheet_obj['columns']
        coltypes = sheet_obj['coltypes']
        colsizes = sheet_obj['colsizes']

        max_height = 0
        for c, col in enumerate(columns):
            value = row.get(col)
            if isinstance(value, list):
                value = ','.join([str(x != None and x or '') for x in value])

            coltype = coltypes.get(col)
            if coltype in ('R', 'F', 'N'):
                print('writing float',c)
                self.writeCell(sheet, current_row, c, value, style="float")
            elif coltype in ('L', 'I'):
                self.writeCell(sheet, current_row, c, value, style="int")
            elif coltype=='D':
                self.writeCell(sheet, current_row, c, value, style="date")
            elif coltype=='DH':
                self.writeCell(sheet, current_row, c, value, style="datetime")
            elif coltype=='DHZ':
                # Excel format doesn't manage timezone: writing the UTC (the user timezone is missing here)
                self.writeCell(sheet, current_row, c, value.replace(tzinfo=None) if value else None, style="datetime")
            else:
                value = toText(value, self.locale)
                self.writeCell(sheet, current_row, c, value, style="default")
            
            heigth, width = self.fitwidth(value)
            max_height = max(max_height, heigth)
            colsizes[c] = max(colsizes.get(c, 0), width)
        sheet.row_dimensions[current_row+1].height = max_height

    def fitwidth(self, data, bold=False):
        # return height, width
        if data is None:
            return 5. , 1.

        # Note: if 'data' is a number the final rendering may provide commas and dots
        #       This case is not considered, column width may be too low for an exact fitting
        #       Same problem for date/datetime values
        if six.PY2:
            if not isinstance(data, (str, unicode)):
                data = str(data)
        elif not isinstance(data, str):
            data = str(data)
        
        rows = data.split('\n')
        
        width = (max([len(x) for x in rows]) + 2. )  # float mandatory
        height = len(rows) * 13.  # float mandatory
        return height, width