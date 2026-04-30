# -*- coding: utf-8 -*-
#--------------------------------------------------------------------------
# package       : GenroPy core - see LICENSE for details
# module flatfiles : flat file readers and writers (CSV, XLS, XLSX, XML)
# Copyright (c) : 2004 - 2007 Softwell sas - Milano
# Written by    : Giovanni Porcari, Michele Bertoldi
#                 Saverio Porcari, Francesco Porcari , Francesco Cavazzana
#--------------------------------------------------------------------------
#This library is free software; you can redistribute it and/or
#modify it under the terms of the GNU Lesser General Public
#License as published by the Free Software Foundation; either
#version 2.1 of the License, or (at your option) any later version.

#This library is distributed in the hope that it will be useful,
#but WITHOUT ANY WARRANTY; without even the implied warranty of
#MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
#Lesser General Public License for more details.

#You should have received a copy of the GNU Lesser General Public
#License along with this library; if not, write to the Free Software
#Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA

"""
Flat file readers (CSV, XLS, XLSX, XML) and writers (XLS, XLSX).
"""

import os
import os.path
import csv
import datetime

from gnr.core import logger
from gnr.core.gnrstring import slugify, toText
from gnr.lib.services.storage import StorageNode
from gnr.core.gnrlist import GnrNamedList  # GnrNamedList lives in gnrlist; imported here for readers

try:
    import xlwt
except Exception:
    logger.warning("Missing xlwt: can't export in xls")

try:
    import openpyxl
except Exception:
    logger.exception("Missing openpyxl: can't export in xlsx")


# ===========================================================================
# Base reader classes
# ===========================================================================

class BaseReader(object):
    """Common base for all flat-file readers providing path decomposition and index building.

    *docname* may be a filesystem path (str) or any file-like object.  When a
    file-like is passed its ``.name`` attribute is used for extension detection
    and logging; readers that cannot operate without a real path (e.g. for
    encoding detection) fall back gracefully.
    """

    def __init__(self, docname, **kwargs):
        self.docname = docname
        if isinstance(docname, str):
            self.dirname = os.path.dirname(docname)
            self.basename, self.ext = os.path.splitext(os.path.basename(docname))
        else:
            name = getattr(docname, 'name', '') or ''
            self.dirname = os.path.dirname(name)
            self.basename, self.ext = os.path.splitext(os.path.basename(name))
        self.ext = self.ext.replace('.', '')

    def _build_index(self, headers, context=''):
        """Build column-name→position index, renaming duplicates as name[pos].

        Modifies *headers* in place for duplicates. Logs a warning when duplicates
        are found, using *context* as the location prefix in the message.
        """
        index = {}
        errors = []
        for i, k in enumerate(headers):
            if k in index:
                new_k = f"{k}[{i}]"
                errors.append(f"Duplicate column '{k}' at position {i}, renamed to '{new_k}'")
                headers[i] = new_k
                index[new_k] = i
            else:
                index[k] = i
        if errors:
            logger.warning(context + ': ' + '; '.join(errors))
        return index, errors


class BaseSheetReader(BaseReader):
    """Base for multi-sheet workbook readers (XLS / XLSX).

    Subclasses implement the workbook-specific primitives; this class owns the
    sheet-discovery loop, the common addSheet logic, the sheet-data properties,
    and the __call__ row generator.
    """

    def __init__(self, docname, mainsheet=None, compressEmptyRows=None, allEmptyRows=None, **kwargs):
        super().__init__(docname, **kwargs)
        self.compressEmptyRows = compressEmptyRows
        self.allEmptyRows = allEmptyRows
        self.sheets = {}
        self._open_workbook()
        for sheetname in self._sheet_names():
            self.addSheet(sheetname)
        self.setMainSheet(mainsheet or self._default_mainsheet())

    # --- subclass primitives ---

    def _open_workbook(self):
        raise NotImplementedError

    def _sheet_names(self):
        raise NotImplementedError

    def _default_mainsheet(self):
        raise NotImplementedError

    def _get_raw_sheet(self, sheetname):
        raise NotImplementedError

    def _sheet_by_index(self, idx):
        raise NotImplementedError

    def _sheet_nrows(self, sheet):
        return 0

    def _process_firstline(self, firstline, sheet):
        """Return (headers_list, colindex_dict) from the raw first row values."""
        raise NotImplementedError

    def _sheetlines(self, sheet):
        raise NotImplementedError

    # --- shared sheet management ---

    def setMainSheet(self, sheetname):
        if isinstance(sheetname, int):
            sheetname = self._sheet_by_index(sheetname)
        self.sheet_base_name = sheetname

    def addSheet(self, sheetname):
        sheet = self._get_raw_sheet(sheetname)
        linegen = self._sheetlines(sheet)
        try:
            firstline = next(linegen)
        except StopIteration:
            firstline = []
        headers, colindex = self._process_firstline(firstline, sheet)
        index, errors = self._build_index(headers, f"Sheet '{sheetname}'")
        self.sheets[sheetname] = {
            'sheet': sheet,
            'headers': headers,
            'colindex': colindex,
            'index': index,
            'ncols': len(headers),
            'nrows': self._sheet_nrows(sheet),
            'errors': '; '.join(errors) if errors else None,
            'linegen': linegen,
        }

    # --- sheet-data properties ---

    @property
    def sheet(self):
        return self.sheets[self.sheet_base_name]['sheet']

    @property
    def headers(self):
        return self.sheets[self.sheet_base_name]['headers']

    @property
    def colindex(self):
        return self.sheets[self.sheet_base_name]['colindex']

    @property
    def index(self):
        return self.sheets[self.sheet_base_name]['index']

    @property
    def ncols(self):
        return self.sheets[self.sheet_base_name]['ncols']

    @property
    def nrows(self):
        return self.sheets[self.sheet_base_name]['nrows']

    def __call__(self, sheetname=None):
        s = self.sheets[sheetname or self.sheet_base_name]
        for line in s['linegen']:
            yield GnrNamedList(s['index'], [c for i, c in enumerate(line) if i in s['colindex']])


# ===========================================================================
# Concrete readers
# ===========================================================================

class XlsReader(BaseSheetReader):
    """Reader for XLS (Excel 97-2003) files using xlrd.

    Supports multiple sheets, automatic column header slugification, and
    handling of duplicate column names. Date cells are converted to datetime.

    :param docname: path to the XLS file, or a binary file-like object
    :param mainsheet: name or index of the main sheet (default: first sheet)
    :param compressEmptyRows: if True, compress consecutive empty rows into one
    :param allEmptyRows: if True, yield all empty rows
    """

    def _open_workbook(self):
        import xlrd
        self.XL_CELL_DATE = xlrd.XL_CELL_DATE
        self.xldate_as_tuple = xlrd.xldate_as_tuple
        if isinstance(self.docname, str):
            self.book = xlrd.open_workbook(filename=self.docname)
        else:
            self.book = xlrd.open_workbook(file_contents=self.docname.read())

    def _sheet_names(self):
        return self.book.sheet_names()

    def _default_mainsheet(self):
        return self.book.sheet_names()[0]

    def _get_raw_sheet(self, sheetname):
        return self.book.sheet_by_name(sheetname)

    def _sheet_by_index(self, idx):
        return self.book.sheet_by_index(idx).name

    def _sheet_nrows(self, sheet):
        return max(sheet.nrows - 1, 0)

    def _process_firstline(self, firstline, sheet):
        headers = [slugify(firstline[c], sep='_') for c in range(min(len(firstline), sheet.ncols))] if firstline else []
        # Build colindex before filtering so it preserves original column positions
        colindex = {i: True for i, h in enumerate(headers) if h}
        headers = [h for h in headers if h]
        return headers, colindex

    def _sheetlines(self, sheet):
        last_line_empty = False
        for lineno in range(sheet.nrows):
            line = sheet.row_values(lineno)
            if [elem for elem in line if elem]:
                row_types = sheet.row_types(lineno)
                for i, c in enumerate(line):
                    if row_types[i] == self.XL_CELL_DATE:
                        try:
                            line[i] = datetime.datetime(*self.xldate_as_tuple(c, sheet.book.datemode))
                        except Exception:
                            line[i] = None
                    if line[i] == '':
                        line[i] = None
                last_line_empty = False
                yield line
            elif self.allEmptyRows:
                yield []
            elif self.compressEmptyRows:
                if not last_line_empty:
                    last_line_empty = True
                    yield []


class XlsxReader(BaseSheetReader):
    """Reader for XLSX (Excel 2007+) files using openpyxl.

    Supports multiple sheets, automatic column header slugification, and
    handling of duplicate column names. Uses read-only mode with lazy loading.
    Empty column headers are renamed to 'gnr_emptycol_N'.

    :param docname: path to the XLSX file, or a binary file-like object
    :param mainsheet: name or index of the main sheet (default: active sheet)
    :param compressEmptyRows: if True, compress consecutive empty rows into one
    :param allEmptyRows: if True, yield all empty rows

    Note: Call close() on the workbook when done.
    """

    def _open_workbook(self):
        from openpyxl import load_workbook
        self.book = load_workbook(
            filename=self.docname,
            read_only=True,
            data_only=True,
            keep_links=False,
        )

    def _sheet_names(self):
        return self.book.sheetnames

    def _default_mainsheet(self):
        return self.book.active.title

    def _get_raw_sheet(self, sheetname):
        return self.book[sheetname]

    def _sheet_by_index(self, idx):
        return self.book.worksheets[idx].title

    def _process_firstline(self, firstline, sheet):
        headers = []
        for i, header in enumerate(firstline):
            if not header:
                header = f'gnr_emptycol_{i}'
            header = slugify(header, sep='_')
            headers.append(header)
        colindex = {i: True for i, h in enumerate(headers) if h}
        return headers, colindex

    def _sheetlines(self, sheet):
        last_line_empty = False
        for line in sheet.rows:
            result = []
            empty_flag = True
            for cell in line:
                value = cell.value
                if value:
                    empty_flag = False
                if value == '':
                    result.append(None)
                else:
                    result.append(value)
            if empty_flag:
                if self.allEmptyRows:
                    yield []
                elif self.compressEmptyRows:
                    if not last_line_empty:
                        last_line_empty = True
                        logger.debug('yielding empty row')
                        yield []
            else:
                last_line_empty = False
                yield result


class CsvReader(BaseReader):
    """Reader for CSV files.

    Supports custom delimiters, encoding detection, and automatic handling of
    duplicate column names. The first row is treated as headers.

    :param docname: path to the CSV file, or a text-mode file-like object.
                   When a file-like is passed it is used directly and will NOT
                   be closed automatically after iteration.
    :param dialect: CSV dialect (e.g., 'excel', 'excel-tab')
    :param delimiter: field delimiter character (default: ',')
    :param detect_encoding: if True, automatically detect file encoding using chardet
                            (ignored when *docname* is a file-like object)
    :param encoding: explicit encoding (currently overridden by detect_encoding logic)
    """

    def __init__(self, docname, dialect=None, delimiter=None, detect_encoding=False,
                 encoding=None, **kwargs):
        super().__init__(docname, **kwargs)
        if isinstance(docname, str):
            # FIXME: why an explicit "encoding" parameter for the constructor but ignoring its value?
            encoding = None
            if detect_encoding:
                encoding = self.detect_encoding()
            self.filecsv = open(docname, 'r', encoding=encoding) if encoding else open(docname, 'r')
            self._owns_filecsv = True
        else:
            self.filecsv = docname
            self._owns_filecsv = False
        self.rows = csv.reader(self.filecsv, dialect=dialect, delimiter=delimiter or ',')
        self.headers = next(self.rows)
        name = docname if isinstance(docname, str) else (getattr(docname, 'name', None) or '<file-like>')
        index, _ = self._build_index(self.headers, f"CSV file '{name}'")
        self.index = index
        self.ncols = len(self.headers)

    def __call__(self):
        for r in self.rows:
            yield GnrNamedList(self.index, r)
        if self._owns_filecsv:
            self.filecsv.close()

    def detect_encoding(self):
        if not isinstance(self.docname, str):
            return None
        try:
            import cchardet as chardet  # noqa: F401
        except ImportError:
            try:
                import chardet  # noqa: F401
            except ImportError:
                logger.exception('either cchardet or chardet are required to detect encoding')
                return
        from chardet.universaldetector import UniversalDetector
        detector = UniversalDetector()
        detector.reset()
        with open(self.docname, 'rb') as f:
            for row in f:
                detector.feed(row)
                if detector.done:
                    break
        detector.close()
        return detector.result.get('encoding')


class XmlReader(BaseReader):
    """Reader for XML files that converts them to GnrNamedList objects.

    Parses XML via Bag and extracts rows based on a collection path or row tag.
    Automatically detects the most common tag as row separator if not specified.

    :param docname: path to the XML file, or a file-like object whose content
                   is read once at construction time
    :param collection_path: optional path to the collection in the Bag structure
    :param row_tag: optional tag name to use as row separator
    """

    def __init__(self, docname, collection_path=None, row_tag=None, **kwargs):
        super().__init__(docname, **kwargs)
        from gnr.core.gnrbag import Bag
        source = docname if isinstance(docname, str) else docname.read()
        self.source = Bag(source)
        if collection_path:
            rows = [n.value.asDict(ascii=True) if n.value else n.attr for n in self.source[collection_path]]
        else:
            if not row_tag:
                if len(self.source) == 1:
                    self.source = self.source['#0']
                from collections import Counter
                row_tag = Counter(list(self.source.keys())).most_common()[0][0]
            rows = [n.value.asDict(ascii=True) if n.value else n.attr for n in self.source if n.label == row_tag]
        self.rows = rows
        r0 = rows[0]
        self.headers = list(r0.keys())
        name = docname if isinstance(docname, str) else (getattr(docname, 'name', None) or '<file-like>')
        index, _ = self._build_index(self.headers, f"XML file '{name}'")
        self.index = index
        self.ncols = len(self.headers)

    def __call__(self):
        for r in self.rows:
            # rows are dicts; extract values in header order
            yield GnrNamedList(self.index, [r.get(k) for k in self.headers])


# ===========================================================================
# Reader factory
# ===========================================================================

def getReader(file_path, filetype=None, **kwargs):
    """Return the appropriate reader instance for *file_path*.

    Auto-detects file type from extension, or uses the explicit *filetype*.

    :param file_path: path string **or** a file-like object.  When a file-like
                     is passed its ``.name`` attribute is used for extension
                     detection; supply an explicit *filetype* when the object
                     has no meaningful name.
    :param filetype: optional override — 'excel', 'tab', 'csv_auto', or None
    :param kwargs: passed to the reader constructor

    File type detection:
        - .xls  -> XlsReader
        - .xlsx -> XlsxReader (falls back to XlsReader if openpyxl missing)
        - .xml  -> XmlReader
        - .tab  -> CsvReader with excel-tab dialect
        - others -> CsvReader

    Example:
        >>> reader = getReader('data.csv')
        >>> for row in reader():
        ...     print(row['column_name'])
    """
    if isinstance(file_path, str):
        _, ext = os.path.splitext(file_path)
    else:
        _, ext = os.path.splitext(getattr(file_path, 'name', '') or '')
    if filetype == 'excel' or (not filetype and ext == '.xls'):
        reader = XlsReader(file_path, **kwargs)
    elif not filetype and ext == '.xlsx':
        try:
            import openpyxl  # noqa: F401
            reader = XlsxReader(file_path, **kwargs)
        except ImportError:  # pragma: no cover
            logger.exception("\n**ERROR Missing openpyxl: 'xlsx' import may not work properly\n")
            reader = XlsReader(file_path, **kwargs)
    elif ext == '.xml':
        reader = XmlReader(file_path, **kwargs)
    else:
        dialect = None
        if filetype == 'tab' or ext == '.tab':
            dialect = 'excel-tab'
        elif filetype == 'csv_auto':
            with open(file_path) as csv_test:
                dialect = csv.Sniffer().sniff(csv_test.read(1024))
        reader = CsvReader(file_path, dialect=dialect, **kwargs)
        reader.index = {slugify(k, sep='_'): v for k, v in reader.index.items()}
    return reader


# ===========================================================================
# Legacy read functions (kept for backwards compatibility)
# ===========================================================================

def readTab(doc):
    """Read a tab-delimited file, yielding GnrNamedList rows.

    :param doc: file path or file-like object
    """
    if isinstance(doc, str):
        f = open(doc)
    else:
        f = doc
    txt = f.read()
    txt = txt.replace('\r\n', '\n')
    txt = txt.replace('\r', '\n')
    lines = txt.split('\n')
    txt = None
    u = [line.split('\t') for line in lines]
    headers = u[0]
    rows = u[1:]
    index = dict([(k, i) for i, k in enumerate(headers)])
    ncols = len(headers)
    for row in rows:
        if len(row) == ncols:
            yield GnrNamedList(index, row)
    if isinstance(doc, str):
        f.close()


def readCSV_new(doc):
    """Read a CSV file — done by Jeff.

    :param doc: file path or file-like object
    """
    if isinstance(doc, str):
        f = open(doc)
    else:
        f = doc
    txt = f.read()
    txt = txt.replace('\r\n', '\n')
    txt = txt.replace('\r', '\n')
    txt = txt.replace('\",\"', '\t')
    txt = txt.replace('\"', '')
    txt = txt.replace(',', '\t')
    lines = txt.split('\n')
    txt = None
    u = [line.split('\t') for line in lines]
    headers = u[0]
    rows = u[1:]
    index = dict([(k, i) for i, k in enumerate(headers)])
    ncols = len(headers)
    for row in rows:
        if len(row) == ncols:
            yield GnrNamedList(index, row)
    if isinstance(doc, str):
        f.close()


def readCSV(doc):
    """Read a CSV/tab file.

    :param doc: file path or file-like object
    """
    if isinstance(doc, str):
        f = open(doc)
    else:
        f = doc
    txt = f.read()
    txt = txt.replace('\r\n', '\n')
    txt = txt.replace('\r', '\n')
    lines = txt.split('\n')
    txt = None
    u = [line.split('\t') for line in lines]
    headers = u[0]
    rows = u[1:]
    index = dict([(k, i) for i, k in enumerate(headers)])
    ncols = len(headers)
    for row in rows:
        if len(row) == ncols:
            yield GnrNamedList(index, row)
    if isinstance(doc, str):
        f.close()


def readXLS(doc):
    """Read an XLS file, yielding GnrNamedList rows.

    :param doc: file path or file-like object
    """
    import xlrd
    if isinstance(doc, str):
        filename = doc
        file_contents = None
    else:
        filename = None
        file_contents = doc.read()
    book = xlrd.open_workbook(filename=filename, file_contents=file_contents)
    sheet = book.sheet_by_index(0)
    headers = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
    headers = [h for h in headers if h]
    index = dict([(k, i) for i, k in enumerate(headers)])
    ncols = len(headers)
    for r in range(1, sheet.nrows):
        row = [sheet.cell_value(r, c) for c in range(ncols)]
        yield GnrNamedList(index, row)


# ===========================================================================
# Writers
# ===========================================================================

class BaseXls(object):

    def save(self, filepath=None, autosize=None):
        self.filepath = filepath
        if autosize is None:
            autosize = True
        self.workbookSave(autosize=autosize)

    def getSheet(self, name=None):
        name = name or self.sheet_base_name
        if not self.sheets:
            self.createSheet(self.sheet_base_name, headers=self.sheet_base_headers,
                             columns=self.sheet_base_columns, coltypes=self.sheet_base_coltypes,
                             groups=self.sheet_base_groups)
        return self.sheets[name]

    @property
    def sheet(self):
        return self.getSheet()['sheet']

    @property
    def headers(self):
        return self.getSheet()['headers']

    @property
    def columns(self):
        return self.getSheet()['columns']

    @property
    def colsizes(self):
        return self.getSheet()['colsizes']

    @property
    def coltypes(self):
        return self.getSheet()['coltypes']

    @property
    def current_row(self):
        return self.getSheet()['current_row']

    @property
    def groups(self):
        return self.getSheet()['groups']

    @property
    def filepath(self):
        return self._filepath

    @filepath.setter
    def filepath(self, filepath):
        if filepath is None:
            self._filepath = None
            self.filenode = None
            return
        self.filenode = None
        if isinstance(filepath, StorageNode):
            self.filenode = filepath
            filepath = self.filenode.path
        filepath = f'{os.path.splitext(filepath)[0]}.{self.extension}'
        if self.filenode:
            self.filenode.path = filepath
        self._filepath = filepath

    def composeAll(self, data=None, **kwargs):
        for export_data in data:
            self.compose(export_data)

    def compose(self, data):
        sheet_name = data['name'].replace('/', '')
        self.createSheet(sheet_name, **data['struct'])
        self.writeHeaders(sheet_name=sheet_name)
        for row in data['rows']:
            self.writeRow(row, sheet_name=sheet_name)


class XlsWriter(BaseXls):
    """Writer for XLS (Excel 97-2003) format using xlwt."""

    extension = 'xls'
    content_type = 'application/xls'

    def __init__(self, columns=None, coltypes=None, headers=None, groups=None, filepath=None,
                 sheet_base_name=None, font='Times New Roman', format_float='#,##0.00',
                 format_int='#,##0', locale=None, print_prefs=None):
        self.sheets = {}
        self.filepath = filepath
        self.workbook = xlwt.Workbook()
        self.print_prefs = print_prefs
        if sheet_base_name is not False:
            self.sheet_base_headers = headers
            self.sheet_base_groups = groups
            self.sheet_base_coltypes = coltypes
            self.sheet_base_columns = columns
            self.sheet_base_name = sheet_base_name or os.path.basename(self.filepath)[:31] if self.filepath else '_base_'
        else:
            self.sheet_base_name = False
        self.locale = locale
        self.float_style = xlwt.XFStyle()
        self.float_style.num_format_str = format_float
        self.int_style = xlwt.XFStyle()
        self.int_style.num_format_str = format_int
        self.date_format = xlwt.XFStyle()
        self.date_format.num_format_str = 'dd/mm/yyyy'
        self.datetime_format = xlwt.XFStyle()
        self.datetime_format.num_format_str = 'dd/mm/yyyy h:mm:ss'
        font0 = xlwt.Font()
        font0.name = font
        font0.bold = True
        self.hstyle = xlwt.XFStyle()
        self.hstyle.font = font0

    def __call__(self, data=None, sheet_name=None):
        self.writeHeaders(sheet_name=sheet_name)
        for item in data:
            row = self.rowGetter(item)
            self.writeRow(row, sheet_name=sheet_name)
        self.workbookSave()

    def createSheet(self, sheetname, headers=None, columns=None, coltypes=None, colsizes=None, groups=None):
        colsizes = colsizes or dict()
        self.sheets[sheetname] = {'sheet': self.workbook.add_sheet(sheetname),
                                   'headers': headers, 'columns': columns,
                                   'colsizes': colsizes, 'coltypes': coltypes,
                                   'groups': groups}
        self.sheets[sheetname]['sheet'].panes_frozen = True
        self.sheets[sheetname]['sheet'].horz_split_pos = 1 if not groups else 2

    def rowGetter(self, item):
        return dict(item)

    def writeHeaders(self, sheet_name=None):
        sheet_name = sheet_name or self.sheet_base_name
        sheet_obj = self.getSheet(sheet_name)
        sheet = sheet_obj['sheet']
        headers = sheet_obj['headers']
        colsizes = sheet_obj['colsizes']
        groups = sheet_obj['groups']
        current_row = 0
        if groups:
            group_style = xlwt.easyxf('align: wrap on, vert centre, horiz center')
            for g in groups:
                name = g.get('name')
                start = g.get('start')
                end = g.get('end')
                sheet.write_merge(current_row, current_row, start, end, name, group_style)
            current_row += 1
        for c, header in enumerate(headers):
            sheet.write(current_row, c, header, self.hstyle)
            colsizes[c] = max(colsizes.get(c, 0), self.fitwidth(header))
        sheet_obj['current_row'] = current_row

    def workbookSave(self, **kwargs):
        if self.filenode:
            with self.filenode.open(mode='wb') as outfile:
                self.workbook.save(outfile)
        else:
            self.workbook.save(self.filepath)

    def write(self, sheet_name=None, what=None):
        pass

    def writeRow(self, row, sheet_name=None):
        sheet_name = sheet_name or self.sheet_base_name
        sheet_obj = self.getSheet(sheet_name)
        current_row = sheet_obj['current_row'] + 1
        sheet_obj['current_row'] = current_row
        sheet = sheet_obj['sheet']
        columns = sheet_obj['columns']
        coltypes = sheet_obj['coltypes']
        colsizes = sheet_obj['colsizes']
        for c, col in enumerate(columns):
            value = row.get(col)
            if isinstance(value, list):
                value = ','.join([str(x != None and x or '') for x in value])
            coltype = coltypes.get(col)
            if coltype in ('R', 'F', 'N'):
                sheet.write(current_row, c, value, self.float_style)
            elif coltype in ('L', 'I'):
                sheet.write(current_row, c, value, self.int_style)
            elif coltype == 'D':
                sheet.write(current_row, c, value, self.date_format)
            elif coltype == 'DH':
                sheet.write(current_row, c, value, self.datetime_format)
            else:
                value = toText(value, self.locale)
                sheet.write(current_row, c, value)
            colsizes[c] = max(colsizes.get(c, 0), self.fitwidth(value))

    def fitwidth(self, data, bold=False):
        """Estimate column width for Arial 10."""
        charwidths = {
            '0': 262.637, '1': 262.637, '2': 262.637, '3': 262.637, '4': 262.637, '5': 262.637, '6': 262.637,
            '7': 262.637, '8': 262.637, '9': 262.637, 'a': 262.637, 'b': 262.637, 'c': 262.637, 'd': 262.637,
            'e': 262.637, 'f': 146.015, 'g': 262.637, 'h': 262.637, 'i': 117.096, 'j': 88.178, 'k': 233.244,
            'l': 88.178, 'm': 379.259, 'n': 262.637, 'o': 262.637, 'p': 262.637, 'q': 262.637, 'r': 175.407,
            's': 233.244, 't': 117.096, 'u': 262.637, 'v': 203.852, 'w': 321.422, 'x': 203.852, 'y': 262.637,
            'z': 233.244, 'A': 321.422, 'B': 321.422, 'C': 350.341, 'D': 350.341, 'E': 321.422, 'F': 291.556,
            'G': 350.341, 'H': 321.422, 'I': 146.015, 'J': 262.637, 'K': 321.422, 'L': 262.637, 'M': 379.259,
            'N': 321.422, 'O': 350.341, 'P': 321.422, 'Q': 350.341, 'R': 321.422, 'S': 321.422, 'T': 262.637,
            'U': 321.422, 'V': 321.422, 'W': 496.356, 'X': 321.422, 'Y': 321.422, 'Z': 262.637, ' ': 146.015,
            '!': 146.015, '"': 175.407, '#': 262.637, '$': 262.637, '%': 438.044, '&': 321.422, '\'': 88.178,
            '(': 175.407, ')': 175.407, '*': 203.852, '+': 291.556, ',': 146.015, '-': 175.407, '.': 146.015,
            '/': 146.015, ':': 146.015, ';': 146.015, '<': 291.556, '=': 291.556, '>': 291.556, '?': 262.637,
            '@': 496.356, '[': 146.015, '\\': 146.015, ']': 146.015, '^': 203.852, '_': 262.637, '`': 175.407,
            '{': 175.407, '|': 146.015, '}': 175.407, '~': 291.556}
        units = 220
        try:
            data = str(data)
        except UnicodeEncodeError:
            return max(units * len(data), 700)
        for char in data:
            if char in charwidths:
                units += charwidths[char]
            else:
                units += charwidths['0']
        if bold:
            units *= 1.1
        return max(units, 700)


class XlsxWriter(BaseXls):
    """Writer for XLSX (Excel 2007+) format using openpyxl."""

    extension = 'xlsx'
    content_type = 'application/xlsx'

    def __init__(self, columns=None, coltypes=None, headers=None, groups=None, filepath=None,
                 sheet_base_name=None, font='Times New Roman', format_float='#,##0.00',
                 format_int='#,##0', format_date=None, format_datetime=None, locale=None,
                 print_prefs=None):
        self.sheets = {}
        self.filepath = filepath
        self.workbook = openpyxl.Workbook()
        self.print_prefs = print_prefs
        del self.workbook[self.workbook.sheetnames[0]]
        if sheet_base_name is not False:
            self.sheet_base_name = sheet_base_name or os.path.basename(self.filepath)[:31] if self.filepath else '_base_'
            self.sheet_base_headers = headers
            self.sheet_base_groups = groups
            self.sheet_base_coltypes = coltypes
            self.sheet_base_columns = columns
        else:
            self.sheet_base_name = False
        self.locale = locale
        format_date = format_date or "D MMM YYYY"
        format_datetime = format_datetime or "D MMM YYYY, H:MM:SS"
        self.workbook.add_named_style(openpyxl.styles.NamedStyle('float',
                                font=openpyxl.styles.Font(name=font),
                                number_format=format_float,
                                alignment=openpyxl.styles.Alignment(vertical="top")))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle('int',
                                font=openpyxl.styles.Font(name=font),
                                number_format=format_int,
                                alignment=openpyxl.styles.Alignment(vertical="top")))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle('header',
                                font=openpyxl.styles.Font(name=font, bold=True),
                                alignment=openpyxl.styles.Alignment(horizontal="center", vertical="top")))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle("default",
                                font=openpyxl.styles.Font(name=font),
                                alignment=openpyxl.styles.Alignment(vertical="top")))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle("date",
                                font=openpyxl.styles.Font(name=font),
                                number_format=format_date,
                                alignment=openpyxl.styles.Alignment(vertical="top")))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle("datetime",
                                font=openpyxl.styles.Font(name=font),
                                number_format=format_datetime,
                                alignment=openpyxl.styles.Alignment(vertical="top")))
        self.workbook.add_named_style(openpyxl.styles.NamedStyle("group",
                                font=openpyxl.styles.Font(name=font, bold=True),
                                alignment=openpyxl.styles.Alignment(vertical="center",
                                                                     wrap_text=True,
                                                                     horizontal='center')))

    def createSheet(self, sheetname, headers=None, columns=None, coltypes=None, colsizes=None, groups=None):
        colsizes = colsizes or dict()
        self.sheets[sheetname] = {'sheet': self.workbook.create_sheet(title=sheetname),
                                   'headers': headers, 'columns': columns,
                                   'colsizes': colsizes, 'coltypes': coltypes,
                                   'groups': groups}
        self.sheets[sheetname]['sheet'].panes_frozen = True
        self.sheets[sheetname]['sheet'].horz_split_pos = 1 if not groups else 2

    def __call__(self, data=None, sheet_name=None):
        self.writeHeaders(sheet_name=sheet_name)
        for item in data:
            row = self.rowGetter(item)
            self.writeRow(row, sheet_name=sheet_name)
        self.workbookSave()

    def rowGetter(self, item):
        return dict(item)

    def writeHeaders(self, sheet_name=None):
        sheet_name = sheet_name or self.sheet_base_name
        sheet_obj = self.getSheet(sheet_name)
        sheet = sheet_obj['sheet']
        self.configurePrintSettings(sheet, sheet_name)
        headers = sheet_obj['headers']
        colsizes = sheet_obj['colsizes']
        groups = sheet_obj['groups']
        current_row = 0
        if groups:
            for g in groups:
                name = g.get('name')
                start = g.get('start')
                end = g.get('end')
                self.writeCell(sheet, current_row, current_row + start, name,
                               style='group', end_column=current_row + end)
            current_row += 1
        max_height = 0
        for c, header in enumerate(headers):
            self.writeCell(sheet, current_row, c, header, style="header")
            heigth, width = self.fitwidth(header)
            max_height = max(max_height, heigth)
            colsizes[c] = max(colsizes.get(c, 0), width)
        sheet_obj['current_row'] = current_row
        sheet.row_dimensions[current_row + 1].height = max_height

    def workbookSave(self, autosize=True):
        if autosize:
            for sheetname in self.sheets:
                ws = self.workbook[sheetname]
                for col, size in self.sheets[sheetname]['colsizes'].items():
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col + 1)].width = size
        if self.filenode:
            with self.filenode.open(mode='wb') as outfile:
                self.workbook.save(filename=outfile)
        else:
            self.workbook.save(filename=self.filepath)

    def configurePrintSettings(self, ws, sheet_name=None):
        if not self.print_prefs:
            return
        for k, v in self.print_prefs.items():
            if not v:
                continue
            if k in ws.page_setup.__dict__:
                ws.page_setup.__dict__[k] = v
            elif k in ws.page_margins.__dict__:
                ws.page_margins.__dict__[k] = v
        if self.print_prefs['show_title'] == 'footer':
            ws.HeaderFooter.oddFooter.center.text = sheet_name
            ws.HeaderFooter.evenFooter.center.text = sheet_name
        elif self.print_prefs['show_title'] == 'header':
            ws.HeaderFooter.oddHeader.center.text = sheet_name
            ws.HeaderFooter.evenHeader.center.text = sheet_name

    def writeCell(self, sheet, row, column, value, style=None, end_column=None):
        cell = sheet.cell(row=row + 1, column=column + 1, value=value)
        if style:
            cell.style = style
        if end_column:
            sheet.merge_cells(start_row=row + 1, start_column=column + 1,
                              end_row=row + 1, end_column=end_column + 1)
        return cell

    def writeRow(self, row, sheet_name=None):
        sheet_name = sheet_name or self.sheet_base_name
        sheet_obj = self.getSheet(sheet_name)
        current_row = sheet_obj['current_row'] + 1
        sheet_obj['current_row'] = current_row
        sheet = sheet_obj['sheet']
        columns = sheet_obj['columns']
        coltypes = sheet_obj['coltypes']
        colsizes = sheet_obj['colsizes']
        max_height = 0
        for c, col in enumerate(columns):
            value = row.get(col)
            if isinstance(value, list):
                value = ','.join([str(x != None and x or '') for x in value])
            coltype = coltypes.get(col)
            if coltype in ('R', 'F', 'N'):
                logger.debug('writing float', c)
                self.writeCell(sheet, current_row, c, value, style="float")
            elif coltype in ('L', 'I'):
                self.writeCell(sheet, current_row, c, value, style="int")
            elif coltype == 'D':
                self.writeCell(sheet, current_row, c, value, style="date")
            elif coltype == 'DH':
                self.writeCell(sheet, current_row, c, value, style="datetime")
            elif coltype == 'DHZ' and not isinstance(value, str):
                self.writeCell(sheet, current_row, c,
                               value.replace(tzinfo=None) if value else None, style="datetime")
            else:
                value = toText(value, self.locale)
                self.writeCell(sheet, current_row, c, value, style="default")
            heigth, width = self.fitwidth(value)
            max_height = max(max_height, heigth)
            colsizes[c] = max(colsizes.get(c, 0), width)
        sheet.row_dimensions[current_row + 1].height = max_height

    def fitwidth(self, data, bold=False):
        """Return (height, width) estimate for *data* string."""
        if data is None:
            return 5., 1.
        if not isinstance(data, str):
            data = str(data)
        rows = data.split('\n')
        width = max([len(x) for x in rows]) + 2.
        height = len(rows) * 13.
        return height, width
