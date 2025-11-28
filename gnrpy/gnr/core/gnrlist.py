# -*- coding: utf-8 -*-
#--------------------------------------------------------------------------
# package       : GenroPy core - see LICENSE for details
# module gnrlis : gnr list implementation
# Copyright (c) : 2004 - 2007 Softwell sas - Milano 
# Written by    : Giovanni Porcari, Michele Bertoldi
#                 Saverio Porcari, Francesco Porcari , Francesco Cavazzana
#--------------------------------------------------------------------------
#This library is free software; you can redistribute it and/or
#modify it under the terms of the GNU Lesser General Public
#License as published by the Free Software Foundation; either
#version 2.1 of the License, or (at your option) any later version.

#This library is distributed in the hope that it will be useful,
#but WITHOUT ANY WARRANTY; without even the implied warranty of
#MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
#Lesser General Public License for more details.

#You should have received a copy of the GNU Lesser General Public
#License along with this library; if not, write to the Free Software
#Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA


"""
Some useful operations on lists.
"""
import os.path
from functools import cmp_to_key
import datetime
import csv

from gnr.core import logger
from gnr.core.gnrdecorator import deprecated
from gnr.core.gnrstring import slugify

# FIXME: what's this for?
class FakeList(list):
    pass

def findByAttr(l, **kwargs):
    """Find elements in the ``l`` list having attributes with names and values as
    kwargs items. Return the list's attributes
    
    :param l: the list"""
    result = list(l)
    for k, v in list(kwargs.items()):
        result = [x for x in result if getattr(x, k, None) == v]
    return result

def hGetAttr(obj, attr):
    """Hierarchical get attribute - traverse nested object attributes using dot notation.

    :param obj: the object to get the attribute from
    :param attr: attribute path, can be hierarchical using dots (e.g., 'user.profile.name')
    :return: the attribute value or None if not found
    """
    if obj is None: return None
    if not '.' in attr:
        return getattr(obj, attr, None)
    else:
        curr, next_item = attr.split('.', 1)
        return hGetAttr(getattr(obj, curr, None), next_item)

def sortByItem(l, *args, **kwargs):
    """Sort the list ``l``, filled of objects with dict interface by items with key in ``*args``.
    Return the list
    
    :param l: the list, where items values must be of consistent type
    :param args: a list of keys to sort for. Each key can be reverse sorted by adding ``:d`` to the key.
    :param hkeys: if ``True`` and a key contains ``.``, then it is interpreted as a hierarchical
                  path and sub dict are looked for"""
    def safeCmp(a, b):
        if a is None:
            if b is None:
                return 0
            return -1
        elif b is None:
            return 1
        else:
            return ((a > b) - (a < b))
            
    def hGetItem(obj, attr):
        if obj is None: return None
        curr, next_item = attr.split('.', 1)
        return hGetAttr(obj.get(curr, None), next_item)
            
    criteria = []
    rev = False
    for crit in list(args):
        caseInsensitive = False
        if ':' in crit:
            crit, direction = crit.split(':', 1)
            if direction.endswith('*'):
                direction = direction[0:-1]
                caseInsensitive = True
            if direction.lower() in['d', 'desc', 'descending']:
                rev = not rev
        criteria = [(crit, rev, caseInsensitive)] + criteria
    hkeys = kwargs.get('hkeys', False)
        
    for crit, rev, caseInsensitive in criteria:
        if caseInsensitive:
            if '.' in crit and hkeys:
                cmp_func = lambda a, b: safeCmp((hGetItem(a, crit) or '').lower(), (hGetItem(b, crit) or '').lower())
                l.sort(key= cmp_to_key(cmp_func))
            else:
                cmp_func = lambda a, b: safeCmp((a.get(crit, None) or '').lower(), (b.get(crit, None) or '').lower())
                l.sort(key= cmp_to_key(cmp_func))
        else:
            if '.' in crit and hkeys:
                cmp_func = lambda a, b: safeCmp(hGetItem(a, crit), hGetItem(b, crit))
                l.sort(key= cmp_to_key(cmp_func))
            else:
                cmp_func = lambda a, b: safeCmp(a.get(crit, None), b.get(crit, None))
                l.sort(key= cmp_to_key(cmp_func))
        if(rev):
            l.reverse()
    return l
        
def sortByAttr(l, *args):
    """Sort a list of objects by their attributes.

    :param l: the list of objects to sort
    :param args: attribute names to sort by. Can include ':d' suffix for descending order.
                 Supports hierarchical attributes with dot notation (e.g., 'user.name')
    :return: the sorted list

    Example:
        sortByAttr(objects, 'name')  # sort by name ascending
        sortByAttr(objects, 'age:d')  # sort by age descending
        sortByAttr(objects, 'dept.name', 'salary:d')  # multi-level sort
    """
    criteria = list(args)
    criteria.reverse()
    for crit in criteria:
        rev = None
        if ':' in crit: crit, rev = crit.split(':', 1)
        if '.' in crit:
            l.sort(key=lambda i:hGetAttr(i, crit))
        else:
            l.sort(key=lambda i:getattr(i, crit, None))
        if rev:
            l.reverse()
    return l

def merge(*args):
    """Merge multiple iterables into a single list, removing duplicates while preserving order.

    Elements from the first iterable are added first, followed by unique elements from
    subsequent iterables.

    :param args: variable number of iterables to merge
    :return: merged list with unique elements

    Note: Elements must be iterable. No type checking is performed.

    Example:
        merge([1, 2, 3], [2, 3, 4], [4, 5])  # returns [1, 2, 3, 4, 5]
    """
    result = list(args[0])
    for l in args[1:]:
        for el in l:
            if not el in result:
                result.append(el)
    return result
        
def readTab(doc):
    """Read a "tab delimited" file.
    
    The :meth:`readCSV()` method was misnamed (read not only CSV files) but must be left for legacy
    
    :param doc: the file to read
    """
    if isinstance(doc, str):
        f = open(doc)
    else:
        f = doc
        
    txt = f.read()
    txt = txt.replace('\r\n', '\n')
    txt = txt.replace('\r', '\n')
    lines = txt.split('\n')
    txt = None
    u = [line.split('\t') for line in lines]
    headers = u[0]
    rows = u[1:]
    
    index = dict([(k, i) for i, k in enumerate(headers)])
    
    ncols = len(headers)
    for row in rows:
        if len(row) == ncols: # it works only for rows with the same length of header
            yield GnrNamedList(index, row)
            
    if isinstance(doc, str):
        f.close()
        
def readCSV_new(doc):
    """This reads a CSV file - done by Jeff
    
    :param doc: the file to read"""
    if isinstance(doc, str):
        f = open(doc)
    else:
        f = doc
        
    txt = f.read()
    txt = txt.replace('\r\n', '\n')
    txt = txt.replace('\r', '\n')
    txt = txt.replace('\",\"', '\t')
    txt = txt.replace('\"', '')
    txt = txt.replace(',', '\t')
    lines = txt.split('\n')
    txt = None
    u = [line.split('\t') for line in lines]
    headers = u[0]
    rows = u[1:]
    
    index = dict([(k, i) for i, k in enumerate(headers)])
    
    ncols = len(headers)
    for row in rows:
        if len(row) == ncols: # it works only for rows with the same length of header
            yield GnrNamedList(index, row)
            
    if isinstance(doc, str):
        f.close()
        
def readCSV(doc):
    """read a CSV file
    
    :param doc: the file to read"""
    if isinstance(doc, str):
        f = open(doc)
    else:
        f = doc
        
    txt = f.read()
    txt = txt.replace('\r\n', '\n')
    txt = txt.replace('\r', '\n')
    lines = txt.split('\n')
    txt = None
    u = [line.split('\t') for line in lines]
    headers = u[0]
    rows = u[1:]
    index = dict([(k, i) for i, k in enumerate(headers)])
    ncols = len(headers)
    for row in rows:
        if len(row) == ncols: # it works only for rows with the same length of header
            yield GnrNamedList(index, row)
            
    if isinstance(doc, str):
        f.close()
        
def readXLS(doc):
    """Read an XLS file
    
    :param doc: the file to read"""
    import xlrd
    
    if isinstance(doc, str):
        filename = doc
        file_contents = None
    else:
        filename = None
        file_contents = doc.read()
        
    book = xlrd.open_workbook(filename=filename, file_contents=file_contents)
    sheet = book.sheet_by_index(0)
    
    headers = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
    headers = [h for h in headers if h]
    
    index = dict([(k, i) for i, k in enumerate(headers)])
    
    ncols = len(headers)
    for r in range(1, sheet.nrows):
        row = [sheet.cell_value(r, c) for c in range(ncols)]
        yield GnrNamedList(index, row)
        
class XlsReader(object):
    """Reader for XLS (Excel 97-2003) files using xlrd library.

    Supports multiple sheets, automatic column header detection with slugification,
    and handling of duplicate column names. Date cells are automatically converted
    to datetime objects.

    :param docname: path to the XLS file
    :param mainsheet: name or index of the main sheet to use (default: first sheet)
    :param compressEmptyRows: if True, compress consecutive empty rows into one
    :param allEmptyRows: if True, yield all empty rows
    """
    def __init__(self, docname,mainsheet=None,compressEmptyRows=None,allEmptyRows=None,**kwargs):
        import xlrd
        self.XL_CELL_DATE = xlrd.XL_CELL_DATE
        self.xldate_as_tuple = xlrd.xldate_as_tuple
        self.docname = docname
        self.dirname = os.path.dirname(docname)
        self.basename, self.ext = os.path.splitext(os.path.basename(docname))
        self.ext = self.ext.replace('.', '')
        self.book = xlrd.open_workbook(filename=self.docname)
        self.compressEmptyRows = compressEmptyRows
        self.allEmptyRows = allEmptyRows
        self.sheets = {}

        for sheetname in self.book.sheet_names():
            self.addSheet(sheetname)
        mainsheet = mainsheet or self.book.sheet_names()[0]
        self.setMainSheet(mainsheet)

    def setMainSheet(self,sheetname):
        if isinstance(sheetname,int):
            sheetname = self.book.sheet_by_index(sheetname).name
        self.sheet_base_name = sheetname

    def addSheet(self,sheetname):
        sheet = self.book.sheet_by_name(sheetname)
        linegen = self._sheetlines(sheet)
        try:
            firstline = next(linegen)
        except StopIteration:
            firstline = []
        headers = [slugify(firstline[c], sep='_') for c in range(min(len(firstline), sheet.ncols))] if firstline else []
        colindex = dict([(i,True)for i,h in enumerate(headers) if h])
        headers = [h for h in headers if h]
        index = dict()
        errors = []
        for i,k in enumerate(headers):
            if k in index:
                # Rename duplicate column with format name[index]
                new_k = f"{k}[{i}]"
                errors.append(f"Duplicate column '{k}' at position {i}, renamed to '{new_k}'")
                headers[i] = new_k
                index[new_k] = i
            else:
                index[k] = i

        if errors:
            logger.warning(f"Sheet '{sheetname}': " + "; ".join(errors))

        self.sheets[sheetname] = {'sheet': sheet,
                                   'headers':headers,
                                   'colindex':colindex,
                                   'index':index,
                                    'ncols':len(headers),
                                    'nrows': max(sheet.nrows - 1, 0),
                                    'errors':'; '.join(errors) if errors else None,
                                    'linegen':linegen}


    @property
    def sheet(self):
        return self.sheets[self.sheet_base_name]['sheet']

    @property
    def headers(self):
        return self.sheets[self.sheet_base_name]['headers']

    @property
    def colindex(self):
        return self.sheets[self.sheet_base_name]['colindex']

    @property
    def index(self):
        return self.sheets[self.sheet_base_name]['index']

    @property
    def ncols(self):
        return self.sheets[self.sheet_base_name]['ncols']

    @property
    def nrows(self):
        return self.sheets[self.sheet_base_name]['nrows']

    def __call__(self,sheetname=None):
        s = self.sheets[sheetname or self.sheet_base_name]
        for line in s['linegen']:
            #row = [self.sheet.cell_value(r, c) for c in range(self.ncols)]
            yield GnrNamedList(s['index'], [c for i,c in enumerate(line) if i in s['colindex']])
            
    def _sheetlines(self,sheet):
        last_line_empty = False
        for lineno in range(sheet.nrows):
            line = sheet.row_values(lineno)
            if [elem for elem in line if elem]:
                row_types = sheet.row_types(lineno)
                for i,c in enumerate(line):
                    if row_types[i] == self.XL_CELL_DATE:
                        try:
                            line[i] = datetime.datetime(*self.xldate_as_tuple(c,sheet.book.datemode))
                        except:
                            line[i] = None
                    if line[i]=='':
                        line[i] = None
                last_line_empty = False
                yield line 
            elif self.allEmptyRows:
                yield []
            elif self.compressEmptyRows:
                if not last_line_empty:
                    last_line_empty = True
                    yield []


class XlsxReader(object):
    """Reader for XLSX (Excel 2007+) files using openpyxl library.

    Supports multiple sheets, automatic column header detection with slugification,
    and handling of duplicate column names. Empty column headers are automatically
    named as 'gnr_emptycol_N' where N is the column index.

    Uses read-only mode with lazy loading for better performance with large files.
    Formula cells return their last calculated value, not the formula itself.

    :param docname: path to the XLSX file
    :param mainsheet: name or index of the main sheet to use (default: active sheet)
    :param compressEmptyRows: if True, compress consecutive empty rows into one
    :param allEmptyRows: if True, yield all empty rows

    Note: The workbook should be explicitly closed with close() method when done.
    """
    def __init__(self, docname, mainsheet=None, compressEmptyRows=None, allEmptyRows=None, **kwargs):
        from openpyxl import load_workbook
        self.docname = docname
        self.dirname = os.path.dirname(docname)
        self.basename, self.ext = os.path.splitext(os.path.basename(docname))
        self.ext = self.ext.replace('.', '')
        self.book = load_workbook(filename=self.docname,
                                  read_only= True,
                                  data_only=True,
                                  keep_links=False)
            # :param data_only: controls whether cells with formulae have either the formula (default) or the value stored the last time Excel read the sheet

            # Unlike a normal workbook, a read-only workbook will use lazy loading.
            # The workbook must be explicitly closed with the close() method.
        self.compressEmptyRows = compressEmptyRows
        self.allEmptyRows = allEmptyRows
        self.sheets = {}

        for sheetname in self.book.sheetnames:
            self.addSheet(sheetname)
          
        mainsheet_name = self.book.active.title
        self.setMainSheet(mainsheet_name)

    def setMainSheet(self,sheetname):
        if isinstance(sheetname, int):
            sheetname = self.book.worksheets(sheetname).title
        self.sheet_base_name = sheetname

    def addSheet(self,sheetname):
        sheet = self.book[sheetname]
        linegen = self._sheetlines(sheet)
        try:
            firstline = next(linegen)
        except StopIteration:
            firstline = []
        headers = []
        for i,header in enumerate(firstline):
            if not header:
                header = f'gnr_emptycol_{i}'
            header = slugify(header, sep='_')
            headers.append(header)
        colindex = dict([(i,True)for i,h in enumerate(headers) if h])
        index = dict()
        errors = []
        for i,k in enumerate(headers):
            if k in index:
                # Rename duplicate column with format name[index]
                new_k = f"{k}[{i}]"
                errors.append(f"Duplicate column '{k}' at position {i}, renamed to '{new_k}'")
                headers[i] = new_k
                index[new_k] = i
            else:
                index[k] = i

        if errors:
            logger.warning(f"Sheet '{sheetname}': " + "; ".join(errors))

        self.sheets[sheetname] = {'sheet': sheet,
                                   'headers':headers,
                                   'colindex':colindex,
                                   'index':index,
                                    'ncols':len(headers),
                                    'nrows':0, #we dont pre-allocate sheet size
                                    'errors':'; '.join(errors) if errors else None,
                                    'linegen':linegen}

    @property
    def sheet(self):
        return self.sheets[self.sheet_base_name]['sheet']

    @property
    def headers(self):
        return self.sheets[self.sheet_base_name]['headers']

    @property
    def colindex(self):
        return self.sheets[self.sheet_base_name]['colindex']

    @property
    def index(self):
        return self.sheets[self.sheet_base_name]['index']

    @property
    def ncols(self):
        return self.sheets[self.sheet_base_name]['ncols']

    @property
    def nrows(self):
        return self.sheets[self.sheet_base_name]['nrows']

    def __call__(self,sheetname=None):
        s = self.sheets[sheetname or self.sheet_base_name]
        for line in s['linegen']:
            #row = [self.sheet.cell_value(r, c) for c in range(self.ncols)]
            yield GnrNamedList(s['index'], [c for i,c in enumerate(line) if i in s['colindex']])
            
    def _sheetlines(self,sheet):
        """Generate lines from the sheet, handling empty rows according to settings.

        :param sheet: openpyxl worksheet object
        :yield: list of cell values for each row
        """
        last_line_empty = False
        for line in sheet.rows:
            result = []
            empty_flag = True
            for cell in line:
                # cell attributes: is_date, data_type => s:string, n:null and numeric, d:date
                value = cell.value
                if value:
                    # Note: the previous [elem for elem in line if elem] also considers zeros
                    empty_flag = False

                if value=='':
                    result.append(None)
                else:
                    result.append(value)

            if empty_flag:
                # self.allEmptyRows and self.compressEmptyRows cannot both be False
                if self.allEmptyRows:
                    yield []
                elif self.compressEmptyRows:
                    if not last_line_empty:
                        last_line_empty = True
                        logger.debug('yielding empty row')
                        yield []
            else:
                last_line_empty = False
                yield result 


class CsvReader(object):
    """Reader for CSV (Comma-Separated Values) files.

    Supports custom delimiters, encoding detection, and automatic handling of
    duplicate column names. The first row is treated as headers.

    :param docname: path to the CSV file
    :param dialect: CSV dialect (e.g., 'excel', 'excel-tab'). If None, uses default
    :param delimiter: field delimiter character (default: ',')
    :param detect_encoding: if True, automatically detect file encoding using chardet
    :param encoding: explicit encoding to use (currently overridden by detect_encoding logic)

    Note: There's a FIXME - the encoding parameter is currently reset to None.
    """
    def __init__(self, docname,dialect=None,delimiter=None,detect_encoding=False,
                encoding=None,**kwargs):
        self.docname = docname
        self.dirname = os.path.dirname(docname)
        self.basename, self.ext = os.path.splitext(os.path.basename(docname))
        self.ext = self.ext.replace('.', '')

        # FIXME: why an explit "encoding" parameter for the constructor but
        # ignoring its value?
        encoding = None
        
        if detect_encoding and not encoding:
            encoding = self.detect_encoding()
        if encoding:
            self.filecsv = open(docname,'r', encoding=encoding)
        else:
            self.filecsv = open(docname,'r')
        self.rows = csv.reader(self.filecsv,dialect=dialect,delimiter=delimiter or ',')
        self.headers = next(self.rows)

        # Handle duplicate columns
        index = dict()
        errors = []
        for i, k in enumerate(self.headers):
            if k in index:
                # Rename duplicate column with format name[index]
                new_k = f"{k}[{i}]"
                errors.append(f"Duplicate column '{k}' at position {i}, renamed to '{new_k}'")
                self.headers[i] = new_k
                index[new_k] = i
            else:
                index[k] = i

        if errors:
            logger.warning(f"CSV file '{docname}': " + "; ".join(errors))

        self.index = index
        self.ncols = len(self.headers)

    def __call__(self):
        for r in self.rows:
            yield GnrNamedList(self.index, r)
        self.filecsv.close()

    def detect_encoding(self):
        try:
            import cchardet as chardet # noqa: F401
        except ImportError:
            try:
                import chardet # noqa: F401
            except ImportError:
                logger.exception('either cchardet or chardet are required to detect encoding')
                return
        from chardet.universaldetector import UniversalDetector
        detector = UniversalDetector()
        detector.reset()
        with open(self.docname, 'rb') as f:
            for row in f:
                detector.feed(row)
                if detector.done: break
        detector.close()
        return detector.result.get('encoding')


class XmlReader(object):
    """Reader for XML files that converts them to GnrNamedList objects.

    Parses XML files using Bag structure and extracts rows based on either a collection
    path or a row tag. Automatically detects the most common tag as row separator if not specified.

    :param docname: path to the XML file
    :param collection_path: optional path to the collection in the Bag structure
    :param row_tag: optional tag name to use as row separator
    """
    def __init__(self, docname,collection_path=None,row_tag=None,**kwargs):
        from gnr.core.gnrbag import Bag
        self.source = Bag(docname)

        if collection_path:
            rows = [n.value.asDict(ascii=True) if n.value else n.attr for n in self.source[collection_path]]
        else:
            if not row_tag:
                if len(self.source)==1:
                    self.source = self.source['#0']
                from collections import Counter
                row_tag = Counter(list(self.source.keys())).most_common()[0][0]
            rows = [n.value.asDict(ascii=True) if n.value else n.attr for n in self.source if n.label == row_tag]
        self.rows = rows
        r0 = rows[0]
        self.headers = list(r0.keys())

        # Handle duplicate columns
        index = dict()
        errors = []
        for i, k in enumerate(self.headers):
            if k in index:
                # Rename duplicate column with format name[index]
                new_k = f"{k}[{i}]"
                errors.append(f"Duplicate column '{k}' at position {i}, renamed to '{new_k}'")
                self.headers[i] = new_k
                index[new_k] = i
            else:
                index[k] = i

        if errors:
            logger.warning(f"XML file '{docname}': " + "; ".join(errors))

        self.index = index
        self.ncols = len(self.headers)

    def __call__(self):
        """Iterate over rows in the XML file.

        :yield: GnrNamedList objects for each row
        """
        for r in self.rows:
            yield GnrNamedList(self.index, r)
        
            
class GnrNamedList(list):
    """A list-like object that allows access to elements by both index and column name.

    This class combines list and dict-like interfaces, enabling both numeric and named
    access to elements. It's primarily used to represent rows from CSV, Excel, or XML files.

    :param index: dict mapping column names (str) to their position indices (int)
    :param values: optional list of initial values ordered according to the index mapping.
                   If None, initializes with None values for all columns.

    Example:
        >>> index = {'name': 0, 'age': 1, 'city': 2}
        >>> row = GnrNamedList(index, ['Alice', 30, 'NYC'])
        >>> row[0]           # Access by numeric index
        'Alice'
        >>> row['name']      # Access by column name
        'Alice'
        >>> row['email'] = 'alice@example.com'  # Add new column dynamically
        >>> 'name' in row    # Check column existence
        True

    Note: Constructor parameters are not type-checked.
    """
    def __init__(self, index, values=None):
        self._index = index
        if values is None:
            self[:] = [None] * len(index)
        else:
            self[:] = values
            
    def __getitem__(self, x):
        if type(x) == int or type(x) == slice:
            # Numeric index or slice - use list's __getitem__ directly
            return list.__getitem__(self, x)
        else:
            # String key - look up in index
            x = self._index[x]
            try:
                return list.__getitem__(self, x)
            except:
                if x > len(self._index):
                    raise
                
    def __contains__(self, what):
        return what in self._index
        
    #def __getattribute__(self, x):
    #    if type(x) != int:
    #        x = self._index[x]
    #    try:
    #        return list.__getattribute__(self, x)
    #    except:
    #        if x > len(self._index):
    #            raise

    #def __delattr__(self,x):
    #    if type(x) != int:
    #        x = self._index[x]
    #    try:
    #        return list.__delattr__(self, x)
    #    except:
    #        if x > len(self._index):
    #            raise
        
    def __setitem__(self, x, v):
        if type(x) not in (int,slice):
            n = self._index.get(x)
            if n is None:
                n = len(self._index)
                self._index[x] = n
            x = n
        try:
            list.__setitem__(self, x, v)
        except:
            n = len(self._index)
            if x > n:
                raise
            else:
                self.extend([None] * (n - len(self)))
                list.__setitem__(self, x, v)
                
    def __str__(self):
        return '[%s]' % ','.join(['%s=%s' % (k, v) for k, v in list(self.items())])
        
    def __repr__(self):
        return '[%s]' % ','.join(['%s=%s' % (k, v) for k, v in list(self.items())])
        
    def get(self, x, default=None):
        """Get value by column name or index, similar to dict.get().

        :param x: column name (str) or numeric index (int)
        :param default: value to return if key/index is not found or raises an exception
        :return: the value at the specified position or default value
        """
        try:
            return self[x]
        except:
            return default
            
    def has_key(self, x):
        """Same of ``has_key`` method's dict. Return ``True`` if the key is in the dict,
        ``False`` otherwise
        
        :param x: the key to test"""
        return x in self._index
        
    def items(self):
        """Same of ``items`` method's dict"""
        items = list(self._index.items())
        result = [None] * len(items)
        for k, v in items:
            result[v] = (k, self[v])
        return result
        
    def iteritems(self):
        """Same of ``iteritems`` method's dict"""
        items = list(self._index.items())
        result = [None] * len(items)
        for k, v in items:
            yield (k, self[v])
            
    def keys(self):
        """Same of ``keys`` method's dict"""
        items = list(self._index.items())
        result = [None] * len(items)
        for k, v in items:
            result[v] = k
        return result
        
    @deprecated(message='do not use pop in named tuple')
    def pop(self, x, dflt=None):
        """Remove and return value at column name or index, similar to dict.pop().

        .. deprecated::
            Do not use pop() with GnrNamedList as it breaks the column index mapping.

        :param x: column name (str) or numeric index (int)
        :param dflt: default value if key not found (currently not used)
        :return: the removed value
        :raises IndexError: if index is out of range
        """
        if type(x) != int:
            x = self._index[x]
        try:
            return list.pop(self, x)
        except:
            if x > len(self._index):
                raise
                
    def update(self, d):
        """Update values from a dict, similar to dict.update().

        :param d: dictionary with column names as keys and new values
        """
        for k, v in list(d.items()):
            self[k] = v
            
    def values(self):
        """Same of ``values`` method's dict"""
        return tuple(self[:] + [None] * (len(self._index) - len(self)))
        
    def extractItems(self, columns):
        """Extract (key, value) pairs for specified columns.

        This is a utility method used by :meth:`fetch() <gnr.sql.gnrsqldata.SqlQuery.fetch()>`.

        :param columns: list of column names to extract. If None or empty, extracts all items
        :return: list of (column_name, value) tuples
        """
        if columns:
            return [(k, self[k]) for k in columns]
        else:
            return list(self.items())

    def extractValues(self, columns):
        """Extract values for specified columns.

        This is a utility method used by :meth:`fetch() <gnr.sql.gnrsqldata.SqlQuery.fetch()>`.

        :param columns: list of column names to extract. If None or empty, extracts all values
        :return: list of values in the order specified by columns
        """
        if columns:
            return [self[k] for k in columns]
        else:
            return list(self.values())     


def getReader(file_path, filetype=None, **kwargs):
    """Factory function to create the appropriate reader based on file type.

    Automatically detects the file type from extension or uses explicit filetype parameter.
    Column names are slugified for CSV/TAB files to ensure valid identifiers.

    :param file_path: path to the file to read
    :param filetype: explicit file type override. Options:
                     - 'excel': force Excel reader (XLS/XLSX)
                     - 'tab': tab-delimited file
                     - 'csv_auto': CSV with automatic dialect detection
                     - None: auto-detect from file extension
    :param kwargs: additional arguments passed to the specific reader constructor

    :return: appropriate reader instance (XlsReader, XlsxReader, CsvReader, or XmlReader)

    File type detection:
        - .xls  -> XlsReader
        - .xlsx -> XlsxReader (falls back to XlsReader if openpyxl not available)
        - .xml  -> XmlReader
        - .tab  -> CsvReader with excel-tab dialect
        - others -> CsvReader

    Example:
        >>> reader = getReader('data.csv')
        >>> for row in reader():
        ...     print(row['column_name'])
    """
    filename, ext = os.path.splitext(file_path)
    if filetype == 'excel' or not filetype and ext in ('.xls','.xlsx'):
        if ext == '.xls':
            reader = XlsReader(file_path,**kwargs)
        else: # .xlsx
            try:
                import openpyxl # noqa: F401
                reader = XlsxReader(file_path,**kwargs)
            except ImportError: # pragma: no cover
                logger.exception("\n**ERROR Missing openpyxl: 'xlsx' import may not work properly\n")
                reader = XlsReader(file_path,**kwargs)
    elif ext=='.xml':
        reader = XmlReader(file_path,**kwargs)
    else:
        dialect = None
        if filetype=='tab' or ext=='.tab':
            dialect = 'excel-tab'
        elif filetype == 'csv_auto':
            with open(file_path) as csv_test:
                dialect = csv.Sniffer().sniff(csv_test.read(1024))

        reader = CsvReader(file_path,dialect=dialect,**kwargs)
        reader.index = {slugify(k, sep='_'):v for k,v in reader.index.items()}
    return reader
