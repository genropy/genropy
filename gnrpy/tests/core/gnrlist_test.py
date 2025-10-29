import os.path
import datetime

import pytest
from gnr.core import gnrlist as gl

def test_findByAttr():
    class MockObj(object):
        pass

    a = MockObj()
    a.name = "colin"
    a.surname = "adams"

    b = MockObj()
    b.name = "eddie"
    b.surname = "adams"
    
    c = MockObj()
    c.name = "arthur"
    c.surname = "dent"

    items = [a,b,c]
    assert a in gl.findByAttr(items, name="colin")
    assert b in gl.findByAttr(items, name="eddie")
    assert c in gl.findByAttr(items, name="arthur")
    assert a in gl.findByAttr(items, name="colin", surname="adams")
    assert b not in gl.findByAttr(items, name="colin", surname="adams")
    assert gl.findByAttr(items, surname="adams") == [a,b]
    assert not gl.findByAttr(items, name="ford")
    
def test_merge():
    merged = gl.merge("foobar", "goober")
    assert merged.count("o") == 2
    assert merged.count("b") == 1
    assert merged.count("e") == 1
    assert merged.count("r") == 1
    

def test_GnrNamedList():
    gnl = gl.GnrNamedList(dict(name=0, surname=1),
                          ["Arthur", "Dent"])

    assert gnl['name'] == "Arthur"
    assert gnl.keys() == ['name','surname']
    for x in gnl.iteritems():
        assert x[0] in ("name", "surname")
        assert x[1] in ("Arthur", "Dent")

    i = gnl.items()
    assert ('name', 'Arthur') in i
    assert ('surname', 'Dent') in i
    assert ('name', 'Ford') not in i
    assert ('surname', 'Prefect') not in i

    assert "name" in gnl
    assert "surname" in gnl
    assert "planet" not in gnl
    
    assert gnl.has_key("name")
    assert gnl.has_key("surname")
    assert not gnl.has_key("planet")

    assert gnl.get("name") == "Arthur"
    assert gnl.get("planet", "Earth") == "Earth"

    assert "name=" in str(gnl)
    assert "surname=" in str(gnl)
    assert "name=" in repr(gnl)
    assert "surname=" in repr(gnl)


    gnl['planet'] = "Earth"
    assert gnl.get('planet') == "Earth"
    
        
    with pytest.raises(IndexError) as excinfo:
        gnl[12] = "goober"
    assert str(excinfo.value) == "list assignment index out of range"

    with pytest.raises(IndexError) as excinfo:
        gnl[122]
    assert str(excinfo.value) == "list index out of range"


    gnl = gl.GnrNamedList(dict(name=0, surname=1))


def test_sortByItem():
    test_l = [
        dict(name="name1",
             surname="surname4",
             age=100,
             company=None,
             birth=datetime.date(2023,3,28)
             ),
        dict(name="name3",
             surname="surname3",
             age=30,
             company=dict(name="ACME, Inc.", address="Via Lemani Dalnaso"),
             birth=datetime.date(2004,3,28)
             ),
        dict(name="name2",
             surname="surname2",
             age=None,
             company={"name":"Wayne Enterprises",
                      "address": {"city":"Gotham"} },
             birth=datetime.date(2004,1,18)
             ),
        dict(name="name2",
             surname="surname1",
             age=20,
             company=None,
             birth=datetime.date(2024,3,28)
             ),
    ]

    res = gl.sortByItem(test_l)

    assert res == test_l
        
    res = gl.sortByItem(test_l, "name:*", hkeys=True)
    assert res[-1]['name'] == "name3"
    res = gl.sortByItem(test_l, "name:d", hkeys=True)
    assert res[-1]['name'] == "name1"
    res = gl.sortByItem(test_l, "name:a", hkeys=True)
    assert res[0]['name'] == "name1"
    res = gl.sortByItem(test_l, "name:a", "surname:d", hkeys=True)
    
    assert res[1]['name'] == res[2]['name'] == 'name2'
    assert res[1]['surname'] == "surname2"
    assert res[2]['surname'] == "surname1"
    res = gl.sortByItem(test_l, "name:a", "surname:a", hkeys=True)
    assert res[1]['name'] == res[2]['name'] == 'name2'
    assert res[1]['surname'] == "surname1"
    assert res[2]['surname'] == "surname2"
    res = gl.sortByItem(test_l, "age")
    assert res[-1]['age'] == 100
    res = gl.sortByItem(test_l, "age:d")
    assert res[0]['age'] == 100

    with pytest.raises(Exception):
        # we can't sort values as dict
        res = gl.sortByItem(test_l, "company", hkeys=True)

    res = gl.sortByItem(test_l, "birth", hkeys=True)
    assert res[0]['name'] == 'name2'
    assert res[0]['birth'] == datetime.date(2004,1,18)
    
    res = gl.sortByItem(test_l, "company.address.city", hkeys=True)
    assert "Wayne" in res[0]['company']['name']

    res = gl.sortByItem(test_l, "company.name:d", hkeys=True)
    assert "Wayne" in res[-1]['company']['name']
    res = gl.sortByItem(test_l, "company.name:d*", hkeys=True)
    assert "Wayne" in res[0]['company']['name']
    
def test_getReader():
    import tempfile
    with tempfile.TemporaryDirectory() as tmpdir:
        filename = os.path.join(tmpdir, 'test.csv')
        with open(filename, "w") as wfp:
            wfp.write("one,two,three\nfour,five,six")
        a = gl.getReader(filename)
        assert isinstance(a, gl.CsvReader)
        a.filecsv.close()

        filename = os.path.join(tmpdir, 'test.tab')
        with open(filename, "w") as wfp:
            wfp.write("one\ttwo\tthree\nfour\tfive\tsix")
        a = gl.getReader(filename)
        assert isinstance(a, gl.CsvReader)
        a.filecsv.close()
        
        filename = os.path.join(tmpdir, 'test.csv')
        with open(filename, "w") as wfp:
            wfp.write("one\ttwo\tthree\nfour\tfive\tsix")
        a = gl.getReader(filename, filetype="csv_auto")
        assert isinstance(a, gl.CsvReader)
        a.filecsv.close()

        # Will fail with an emtpy file
        filename = os.path.join(tmpdir, 'test.xls')
        with open(filename, "w") as wfp:
            pass
        with pytest.raises(Exception):
            a = gl.getReader(filename, filetype="excel")
            a.filecsv.close()
        filename = os.path.join(tmpdir, 'test.xlsx')
        with open(filename, "w") as wfp:
            pass
        with pytest.raises(Exception):
            a = gl.getReader(filename, filetype="excel")
            a.filecsv.close()
            
    test_dir = os.path.dirname(__file__)
    
    filename = os.path.join(test_dir, "data", "test.xls")
    a = gl.getReader(filename)
    assert isinstance(a, gl.XlsReader)

    filename = os.path.join(test_dir, "data","test.xlsx")
    a = gl.getReader(filename)
    assert isinstance(a, gl.XlsxReader)

    # FIXME: this fails all the time.
    with pytest.raises(Exception):
        filename = os.path.join(test_dir, "data", "testbag.xml")
        a = gl.getReader(filename)

def test_CsvReader():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", "test.csv")
    a = gl.CsvReader(test_file)
    # FIXME: odd interface using __call__
    r = [x for x in a()]
    assert len(r) == 1
    assert isinstance(r[0], gl.GnrNamedList)
    assert 'a' in r[0].keys()
    a = gl.CsvReader(test_file, detect_encoding=True)


def test_XlsReader():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.xls')
    r = gl.XlsReader(test_file)
    assert r.sheet.name == "Sheet1"
    assert 'a' in r.headers
    assert 0 in r.colindex
    assert r.colindex[0] is True
    assert 'a' in r.index
    assert r.ncols == 3
    assert r.nrows == 1
    d = [x for x in r()]
    assert len(d) == 1
    assert isinstance(d[0], gl.GnrNamedList)
    assert 'a' in d[0].keys()

def test_XlsxReader():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.xlsx')
    r = gl.XlsxReader(test_file)
    assert r.sheet.title == "Sheet1"
    assert 'a' in r.headers
    assert 0 in r.colindex
    assert r.colindex[0] is True
    assert 'a' in r.index
    assert r.ncols == 3

    # FIXME: this doesn't work in the implementation
    #assert r.nrows == 1

    d = [x for x in r()]
    assert len(d) == 1
    assert isinstance(d[0], gl.GnrNamedList)
    assert 'a' in d[0].keys()

def test_readXLS():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.xls')
    r = gl.readXLS(test_file)
    d = [x for x in r]
    assert len(d) == 1
    assert isinstance(d[0], gl.GnrNamedList)
    assert 'a' in d[0].keys()

    with open(test_file, "rb") as fp:
        r = gl.readXLS(fp)
        d = [x for x in r]
        assert len(d) == 1
        assert isinstance(d[0], gl.GnrNamedList)
        assert 'a' in d[0].keys()

def test_readCSV():
    # FIXME: apparently, readXLS and readCSV exposes
    # a different interface to access record, please
    # check the last assert here with the last of readXLS test
    # while readCSV_new works correctly. Maybe the _new should
    # be the implementation..
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.csv')
    r = gl.readCSV(test_file)
    d = [x for x in r]
    assert len(d) == 2
    assert isinstance(d[0], gl.GnrNamedList)
    assert 'a' in d[0].keys()[0]

    with open(test_file, "r") as fp:
        r = gl.readCSV(fp)
        d = [x for x in r]
        assert len(d) == 2
        assert isinstance(d[0], gl.GnrNamedList)
        assert 'a' in d[0].keys()[0]

def test_readCSV_new():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.csv')
    r = gl.readCSV_new(test_file)
    d = [x for x in r]
    assert len(d) == 1
    assert isinstance(d[0], gl.GnrNamedList)
    assert 'a' in d[0].keys()

    with open(test_file, "r") as fp:
        r = gl.readCSV_new(fp)
        d = [x for x in r]
        assert len(d) == 1
        assert isinstance(d[0], gl.GnrNamedList)
        assert 'a' in d[0].keys()

        
def test_sortByAttr():
    class MockObj(object):
        a = 1

    m1 = MockObj()
    m2 = MockObj()
    m2.a = 2
    m3 = MockObj()
    m3.a = 3
    test_l = [m1, m2, m3]
    r = gl.sortByAttr(test_l, "a")

    m1 = MockObj()
    m1.a = MockObj()
    m2 = MockObj()
    m2.a = MockObj()


def test_CsvReader_duplicate_columns():
    """Test handling of duplicate column names in CSV files"""
    import tempfile
    import csv

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        csv_file = f.name
        writer = csv.writer(f)
        # Headers with duplicate 'name' column
        writer.writerow(['id', 'name', 'surname', 'name', 'email'])
        writer.writerow(['1', 'Mario', 'Rossi', 'Giuseppe', 'mario@test.com'])
        writer.writerow(['2', 'Laura', 'Bianchi', 'Anna', 'laura@test.com'])

    try:
        reader = gl.CsvReader(csv_file)

        # Check that duplicate column has been renamed
        assert 'name' in reader.headers
        assert 'name[3]' in reader.headers
        assert reader.headers == ['id', 'name', 'surname', 'name[3]', 'email']

        # Check index mapping
        assert reader.index['name'] == 1
        assert reader.index['name[3]'] == 3

        # Read rows
        rows = [row for row in reader()]
        assert len(rows) == 2

        # Test first row
        row = rows[0]
        assert row[0] == '1'
        assert row[1] == 'Mario'
        assert row[2] == 'Rossi'
        assert row[3] == 'Giuseppe'
        assert row[4] == 'mario@test.com'

        # Access by name
        assert row['id'] == '1'
        assert row['name'] == 'Mario'
        assert row['surname'] == 'Rossi'
        assert row['name[3]'] == 'Giuseppe'
        assert row['email'] == 'mario@test.com'

    finally:
        os.unlink(csv_file)


def test_XlsxReader_duplicate_columns():
    """Test handling of duplicate column names in XLSX files"""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    import tempfile

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        xlsx_file = f.name

    try:
        # Create XLSX with duplicate columns
        wb = Workbook()
        ws = wb.active
        ws.append(['id', 'name', 'surname', 'name', 'email'])
        ws.append(['1', 'Mario', 'Rossi', 'Giuseppe', 'mario@test.com'])
        ws.append(['2', 'Laura', 'Bianchi', 'Anna', 'laura@test.com'])
        wb.save(xlsx_file)

        reader = gl.XlsxReader(xlsx_file)

        # Check that duplicate column has been renamed
        assert 'name' in reader.headers
        assert 'name[3]' in reader.headers
        assert reader.headers == ['id', 'name', 'surname', 'name[3]', 'email']

        # Check index mapping
        assert reader.index['name'] == 1
        assert reader.index['name[3]'] == 3

        # Read rows
        rows = [row for row in reader()]
        assert len(rows) == 2

        # Test first row
        row = rows[0]
        assert row[0] == '1'
        assert row[1] == 'Mario'
        assert row[2] == 'Rossi'
        assert row[3] == 'Giuseppe'
        assert row[4] == 'mario@test.com'

        # Access by name
        assert row['id'] == '1'
        assert row['name'] == 'Mario'
        assert row['surname'] == 'Rossi'
        assert row['name[3]'] == 'Giuseppe'
        assert row['email'] == 'mario@test.com'

    finally:
        os.unlink(xlsx_file)


def test_XlsReader_duplicate_columns():
    """Test handling of duplicate column names in XLS files"""
    try:
        import xlwt
    except ImportError:
        pytest.skip("xlwt not available")

    import tempfile

    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as f:
        xls_file = f.name

    try:
        # Create XLS with duplicate columns
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Sheet1')

        # Headers
        headers = ['id', 'name', 'surname', 'name', 'email']
        for col, header in enumerate(headers):
            ws.write(0, col, header)

        # Data rows
        data = [
            ['1', 'Mario', 'Rossi', 'Giuseppe', 'mario@test.com'],
            ['2', 'Laura', 'Bianchi', 'Anna', 'laura@test.com']
        ]
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data):
                ws.write(row_idx, col_idx, value)

        wb.save(xls_file)

        reader = gl.XlsReader(xls_file)

        # Check that duplicate column has been renamed
        assert 'name' in reader.headers
        assert 'name[3]' in reader.headers
        assert reader.headers == ['id', 'name', 'surname', 'name[3]', 'email']

        # Check index mapping
        assert reader.index['name'] == 1
        assert reader.index['name[3]'] == 3

        # Read rows
        rows = [row for row in reader()]
        assert len(rows) == 2

        # Test first row
        row = rows[0]
        assert row[0] == '1'
        assert row[1] == 'Mario'
        assert row[2] == 'Rossi'
        assert row[3] == 'Giuseppe'
        assert row[4] == 'mario@test.com'

        # Access by name
        assert row['id'] == '1'
        assert row['name'] == 'Mario'
        assert row['surname'] == 'Rossi'
        assert row['name[3]'] == 'Giuseppe'
        assert row['email'] == 'mario@test.com'

    finally:
        os.unlink(xls_file)


def test_hGetAttr():
    """Test hierarchical attribute getter"""
    class MockObj:
        def __init__(self):
            self.name = "Alice"
            self.profile = None

    class Profile:
        def __init__(self):
            self.city = "NYC"

    obj = MockObj()
    obj.profile = Profile()

    # Simple attribute
    assert gl.hGetAttr(obj, 'name') == "Alice"

    # Hierarchical attribute
    assert gl.hGetAttr(obj, 'profile.city') == "NYC"

    # Non-existent attribute
    assert gl.hGetAttr(obj, 'nonexistent') is None

    # None object
    assert gl.hGetAttr(None, 'anything') is None

    # Nested None
    obj.profile = None
    assert gl.hGetAttr(obj, 'profile.city') is None


def test_readTab():
    """Test tab-delimited file reading"""
    import tempfile

    with tempfile.NamedTemporaryFile(mode='w', suffix='.tab', delete=False) as f:
        tab_file = f.name
        f.write("name\tage\tcity\n")
        f.write("Alice\t30\tNYC\n")
        f.write("Bob\t25\tLA\n")

    try:
        rows = list(gl.readTab(tab_file))
        assert len(rows) == 2
        assert isinstance(rows[0], gl.GnrNamedList)
        assert rows[0]['name'] == 'Alice'
        assert rows[0]['age'] == '30'
        assert rows[0]['city'] == 'NYC'
        assert rows[1]['name'] == 'Bob'

    finally:
        os.unlink(tab_file)


def test_GnrNamedList_extractMethods():
    """Test extractItems and extractValues methods"""
    index = {'name': 0, 'age': 1, 'city': 2}
    row = gl.GnrNamedList(index, ['Alice', 30, 'NYC'])

    # extractItems with specific columns
    items = row.extractItems(['name', 'city'])
    assert items == [('name', 'Alice'), ('city', 'NYC')]

    # extractItems with all columns
    all_items = row.extractItems(None)
    assert len(all_items) == 3
    assert ('name', 'Alice') in all_items

    # extractValues with specific columns
    values = row.extractValues(['age', 'name'])
    assert values == [30, 'Alice']

    # extractValues with all columns
    all_values = row.extractValues(None)
    assert len(all_values) == 3
    assert 'Alice' in all_values


def test_GnrNamedList_dynamic_columns():
    """Test dynamic column addition"""
    index = {'name': 0, 'age': 1}
    row = gl.GnrNamedList(index, ['Alice', 30])

    # Add new column
    row['city'] = 'NYC'
    assert row['city'] == 'NYC'
    assert 'city' in row
    assert row._index['city'] == 2

    # Update existing column
    row['age'] = 31
    assert row['age'] == 31

    # Add another new column
    row['country'] = 'USA'
    assert row['country'] == 'USA'
    assert len(row._index) == 4


def test_multiple_duplicate_columns():
    """Test handling of 3+ duplicate columns"""
    import tempfile
    import csv

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        csv_file = f.name
        writer = csv.writer(f)
        # Three 'name' columns
        writer.writerow(['id', 'name', 'name', 'name', 'email'])
        writer.writerow(['1', 'First', 'Middle', 'Last', 'test@test.com'])

    try:
        reader = gl.CsvReader(csv_file)

        # Check all duplicate columns are renamed
        assert reader.headers == ['id', 'name', 'name[2]', 'name[3]', 'email']
        assert reader.index['name'] == 1
        assert reader.index['name[2]'] == 2
        assert reader.index['name[3]'] == 3

        rows = list(reader())
        row = rows[0]

        # All values accessible
        assert row['name'] == 'First'
        assert row['name[2]'] == 'Middle'
        assert row['name[3]'] == 'Last'

    finally:
        os.unlink(csv_file)


def test_GnrNamedList_iteritems():
    """Test iteritems method"""
    index = {'name': 0, 'age': 1, 'city': 2}
    row = gl.GnrNamedList(index, ['Alice', 30, 'NYC'])

    items_list = list(row.iteritems())
    assert len(items_list) == 3
    assert ('name', 'Alice') in items_list
    assert ('age', 30) in items_list
    assert ('city', 'NYC') in items_list


def test_GnrNamedList_values():
    """Test values method"""
    index = {'name': 0, 'age': 1}
    row = gl.GnrNamedList(index, ['Alice', 30])

    values = row.values()
    assert isinstance(values, tuple)
    assert len(values) == 2
    assert 'Alice' in values
    assert 30 in values


def test_sortByItem_case_insensitive():
    """Test case-insensitive sorting"""
    test_l = [
        {'name': 'alice', 'age': 30},
        {'name': 'Charlie', 'age': 25},
        {'name': 'bob', 'age': 35},
    ]

    # Case-insensitive sort
    result = gl.sortByItem(test_l, 'name:a*')
    assert result[0]['name'] == 'alice'
    assert result[1]['name'] == 'bob'
    assert result[2]['name'] == 'Charlie'


def test_slugify_consistency_across_readers():
    """Test that CSV, XLS, and XLSX readers use consistent slugification"""
    import tempfile
    import csv

    # Headers with spaces and special characters
    headers = ['Transaction ID', 'User Name', 'Email-Address', 'Created At']
    expected_keys = ['transaction_id', 'user_name', 'email_address', 'created_at']

    # Test CSV via getReader
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        csv_file = f.name
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerow(['123', 'Alice', 'alice@test.com', '2025-10-28'])

    try:
        csv_reader = gl.getReader(csv_file)
        csv_keys = list(csv_reader.index.keys())
        assert csv_keys == expected_keys, f"CSV keys: {csv_keys}"

        # Verify access works with underscores
        for row in csv_reader():
            assert row['transaction_id'] == '123'
            assert row['user_name'] == 'Alice'
            break
    finally:
        os.unlink(csv_file)

    # Test XLSX
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        xlsx_file = f.name

    try:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(['123', 'Alice', 'alice@test.com', '2025-10-28'])
        wb.save(xlsx_file)

        xlsx_reader = gl.XlsxReader(xlsx_file)
        xlsx_keys = list(xlsx_reader.index.keys())
        assert xlsx_keys == expected_keys, f"XLSX keys: {xlsx_keys}"

        # Verify access works with underscores
        for row in xlsx_reader():
            assert row['transaction_id'] == '123'
            assert row['user_name'] == 'Alice'
            break
    finally:
        os.unlink(xlsx_file)

    # Test XLS
    try:
        import xlwt
    except ImportError:
        pytest.skip("xlwt not available")

    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as f:
        xls_file = f.name

    try:
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Sheet1')
        for col, header in enumerate(headers):
            ws.write(0, col, header)
        for col, value in enumerate(['123', 'Alice', 'alice@test.com', '2025-10-28']):
            ws.write(1, col, value)
        wb.save(xls_file)

        xls_reader = gl.XlsReader(xls_file)
        xls_keys = list(xls_reader.index.keys())
        assert xls_keys == expected_keys, f"XLS keys: {xls_keys}"

        # Verify access works with underscores
        for row in xls_reader():
            assert row['transaction_id'] == '123'
            assert row['user_name'] == 'Alice'
            break
    finally:
        os.unlink(xls_file)


def test_GnrNamedList_sql_adapter_compatibility():
    """Test GnrNamedList works correctly when used by SQL adapter (gnrdict_row)"""
    # Simulate how gnrdict_row creates GnrNamedList
    index = {'id': 0, 'name': 1, 'email': 2}
    values = [1, 'John Doe', 'john@example.com']

    row = gl.GnrNamedList(index, values=values)

    # Test numeric access (as list)
    assert row[0] == 1
    assert row[1] == 'John Doe'
    assert row[2] == 'john@example.com'

    # Test named access (as dict)
    assert row['id'] == 1
    assert row['name'] == 'John Doe'
    assert row['email'] == 'john@example.com'

    # Test slice access (critical for SQL adapter)
    assert row[:] == [1, 'John Doe', 'john@example.com']
    assert row[0:2] == [1, 'John Doe']
    assert row[1:] == ['John Doe', 'john@example.com']

    # Test negative indexing
    assert row[-1] == 'john@example.com'
    assert row[-2] == 'John Doe'

    # Test iteration
    result = list(row)
    assert result == [1, 'John Doe', 'john@example.com']

    # Test len
    assert len(row) == 3


def test_GnrNamedList_sql_adapter_with_duplicates():
    """Test GnrNamedList with duplicate column names (from reader with duplicates)"""
    # Simulate reader that renamed duplicate columns
    index = {'id': 0, 'value': 1, 'value[2]': 2, 'value[3]': 3}
    values = [1, 10, 20, 30]

    row = gl.GnrNamedList(index, values=values)

    # Test all columns are accessible
    assert row['id'] == 1
    assert row['value'] == 10
    assert row['value[2]'] == 20
    assert row['value[3]'] == 30

    # Test numeric access still works
    assert row[0] == 1
    assert row[1] == 10
    assert row[2] == 20
    assert row[3] == 30

    # Test slice access
    assert row[:] == [1, 10, 20, 30]
    assert row[1:] == [10, 20, 30]


def test_GnrNamedList_mixed_access_patterns():
    """Test GnrNamedList with various access patterns used in production code"""
    index = {'transaction_id': 0, 'user_name': 1, 'amount': 2}
    values = ['TX123', 'Alice', 100.50]

    row = gl.GnrNamedList(index, values=values)

    # Common patterns used in tableImporterCheck and related code

    # Pattern 1: Iterate and access by name
    for item in row:
        assert item is not None

    # Pattern 2: Convert to dict
    row_dict = dict(row)
    assert 'transaction_id' in row_dict or 0 in row_dict

    # Pattern 3: Slice and process
    subset = row[1:]
    assert len(subset) == 2
    assert subset[0] == 'Alice'

    # Pattern 4: Check membership (for index)
    assert 'transaction_id' in row._index
    assert 'user_name' in row._index

    # Pattern 5: Enumerate
    for i, value in enumerate(row):
        assert row[i] == value

