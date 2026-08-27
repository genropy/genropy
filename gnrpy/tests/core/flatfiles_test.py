# -*- coding: utf-8 -*-
"""
Tests for gnr.core.flatfiles.

These tests import directly from flatfiles (no deprecation warnings) and cover:
  - BaseReader path decomposition with file-like objects
  - File-like input for CsvReader, XlsReader, XlsxReader, XmlReader
  - CsvReader file-ownership semantics (open/close)
  - CsvReader.detect_encoding
  - XmlReader with properly tabular XML
  - getReader dispatch for both paths and file-like objects
  - XlsWriter and XlsxWriter round-trips (write then read back)
"""
import io
import os
import tempfile
import csv

import pytest

from gnr.core.flatfiles import (
    CsvReader, XlsReader, XlsxReader, XmlReader,
    getReader, getCsvDialect, readTab, readCSV, readCSV_new, readXLS,
    XlsWriter, XlsxWriter,
)
from gnr.core.gnrlist import GnrNamedList

TEST_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(TEST_DIR, 'data')

# Minimal tabular XML: two rows, attribute-style values
_TABULAR_XML = """\
<?xml version="1.0" encoding="UTF-8"?>
<GenRoBag>
  <row><name>Alice</name><city>NYC</city></row>
  <row><name>Bob</name><city>LA</city></row>
</GenRoBag>
"""


# ===========================================================================
# Direct imports from flatfiles (smoke: no DeprecationWarning)
# ===========================================================================

def test_direct_import_no_warning(recwarn):
    """Importing from flatfiles must not raise DeprecationWarning."""
    from gnr.core.flatfiles import CsvReader as CR  # noqa: F401
    dep_warnings = [w for w in recwarn.list if issubclass(w.category, DeprecationWarning)]
    assert dep_warnings == []


# ===========================================================================
# BaseReader — path decomposition for file-like objects
# ===========================================================================

def test_basereader_filelike_with_name():
    """ext/basename/dirname are derived from .name when the input is a file-like."""
    f = io.StringIO("a,b\n1,2")
    f.name = '/tmp/report.csv'
    r = CsvReader(f)
    assert r.ext == 'csv'
    assert r.basename == 'report'
    assert r.dirname == '/tmp'


def test_basereader_filelike_without_name():
    """File-like with no .name produces empty path components, no crash."""
    f = io.StringIO("a,b\n1,2")
    r = CsvReader(f)
    assert r.ext == ''
    assert r.basename == ''
    assert r.dirname == ''


# ===========================================================================
# CsvReader — file-like support and ownership semantics
# ===========================================================================

def test_CsvReader_stringio_reads_correctly():
    """CsvReader accepts a StringIO and yields correct rows."""
    f = io.StringIO("name,age\nAlice,30\nBob,25")
    r = CsvReader(f)
    assert r.headers == ['name', 'age']
    assert r.ncols == 2
    rows = list(r())
    assert len(rows) == 2
    assert isinstance(rows[0], GnrNamedList)
    assert rows[0]['name'] == 'Alice'
    assert rows[1]['age'] == '25'


def test_CsvReader_filelike_not_closed_after_iteration():
    """Reader does not close a caller-provided file-like after iteration."""
    f = io.StringIO("x,y\n1,2")
    r = CsvReader(f)
    assert r._owns_filecsv is False
    list(r())
    assert not f.closed


def test_CsvReader_path_owns_and_closes_file():
    """Reader opened from a path sets _owns_filecsv and closes after iteration."""
    path = os.path.join(DATA_DIR, 'test.csv')
    r = CsvReader(path)
    assert r._owns_filecsv is True
    list(r())
    assert r.filecsv.closed


def test_CsvReader_filelike_with_name_attribute():
    """Context string in duplicate-column warnings uses .name when available."""
    f = io.StringIO("col,col\n1,2")
    f.name = 'upload.csv'
    r = CsvReader(f)   # duplicate column 'col' triggers a warning
    assert 'col[1]' in r.headers
    assert r.ncols == 2


# ===========================================================================
# CsvReader.detect_encoding
# ===========================================================================

def test_CsvReader_detect_encoding_path():
    """detect_encoding works for common file encodings."""
    pytest.importorskip('chardet')
    for filename in ('test_Enc_UTF8.csv', 'test_Enc_ISO8859_1.csv'):
        path = os.path.join(DATA_DIR, filename)
        r = CsvReader(path, detect_encoding=True)
        rows = list(r())
        assert len(rows) > 0


def test_CsvReader_detect_encoding_filelike_is_noop():
    """detect_encoding is silently skipped for file-like inputs."""
    f = io.StringIO("a,b\n1,2")
    r = CsvReader(f, detect_encoding=True)
    rows = list(r())
    assert len(rows) == 1


def test_CsvReader_encoding_detection():
    """Test CsvReader with detect_encoding=True on files with various encodings.

    Verifies that:
    1. Header is correctly read (slugified to 'id,nome,citta,descrizione' or similar)
    2. Last row, 4th field (descrizione) contains encoding-specific characters
    """
    # Map filename to expected last row description field
    # Each file has different content in the last row's 'descrizione' field
    test_cases = [
        ('test_Enc_ASCII.csv', ['id', 'nome', 'citta', 'descrizione'], 'Descrizione semplice e pulita'),
        ('test_Enc_UTF8.csv', ['id', 'nome', 'città', 'descrizione'], 'Текст на русском языке'),
        ('test_Enc_ISO8859_1.csv', ['id', 'nome', 'città', 'descrizione'], 'Größe und Qualität'),
        ('test_Enc_Windows1252.csv', ['id', 'nome', 'città', 'descrizione'], 'Größe™ und Qualität'),
        ('test_Enc_Windows1251.csv', ['id', 'name', 'city', 'description'], 'Белорусская столица'),
        ('test_Enc_Windows1253.csv', ['id', 'name', 'city', 'description'], 'Νησί της Κρήτης €'),
        ('test_Enc_KOI8R.csv', ['id', 'name', 'city', 'description'], 'Дальний Восток России'),
        ('test_Enc_SHIFTJIS.csv', ['id', 'name', 'city', 'description'], '北海道の首都'),
        ('test_Enc_EUCKR.csv', ['id', 'name', 'city', 'description'], '영남지방 중심도시'),
        ('test_Enc_GB2312.csv', ['id', 'nome', 'città', 'descrizione'], '经济特区城市'),
    ]

    for filename, expected_header, expected_last_descrizione in test_cases:
        test_file = os.path.join(DATA_DIR, filename)

        reader = CsvReader(test_file, detect_encoding=True)

        assert reader.headers == expected_header

        rows = list(reader())
        last_row = rows[-1]

        assert last_row[3] == expected_last_descrizione


# ===========================================================================
# XlsReader — file-like support
# ===========================================================================

def test_XlsReader_bytesio():
    """XlsReader reads from a BytesIO (binary file-like)."""
    xls_path = os.path.join(DATA_DIR, 'test.xls')
    with open(xls_path, 'rb') as f:
        data = io.BytesIO(f.read())
    r = XlsReader(data)
    assert 'a' in r.headers
    rows = list(r())
    assert len(rows) == 1
    assert isinstance(rows[0], GnrNamedList)
    assert 'a' in rows[0]


def test_XlsReader_bytesio_with_name():
    """XlsReader resolves ext/basename from BytesIO.name."""
    xls_path = os.path.join(DATA_DIR, 'test.xls')
    with open(xls_path, 'rb') as f:
        data = io.BytesIO(f.read())
    data.name = 'report.xls'
    r = XlsReader(data)
    assert r.ext == 'xls'
    assert r.basename == 'report'
    rows = list(r())
    assert len(rows) == 1


# ===========================================================================
# XlsxReader — file-like support
# ===========================================================================

def test_XlsxReader_bytesio():
    """XlsxReader reads from a BytesIO (openpyxl accepts file-like natively)."""
    pytest.importorskip('openpyxl')
    xlsx_path = os.path.join(DATA_DIR, 'test.xlsx')
    with open(xlsx_path, 'rb') as f:
        data = io.BytesIO(f.read())
    r = XlsxReader(data)
    assert 'a' in r.headers
    rows = list(r())
    assert len(rows) == 1
    assert isinstance(rows[0], GnrNamedList)
    assert 'a' in rows[0]


def test_XlsxReader_bytesio_with_name():
    """XlsxReader resolves ext/basename from BytesIO.name."""
    pytest.importorskip('openpyxl')
    xlsx_path = os.path.join(DATA_DIR, 'test.xlsx')
    with open(xlsx_path, 'rb') as f:
        data = io.BytesIO(f.read())
    data.name = 'report.xlsx'
    r = XlsxReader(data)
    assert r.ext == 'xlsx'
    assert r.basename == 'report'


# ===========================================================================
# XmlReader — tabular XML and file-like support
# ===========================================================================

def test_XmlReader_path_explicit_row_tag(tmp_path):
    """XmlReader reads a tabular XML file with an explicit row_tag."""
    xml_file = tmp_path / 'data.xml'
    xml_file.write_text(_TABULAR_XML)
    r = XmlReader(str(xml_file), row_tag='row')
    assert 'name' in r.headers
    assert 'city' in r.headers
    rows = list(r())
    assert len(rows) == 2
    assert isinstance(rows[0], GnrNamedList)
    assert rows[0]['name'] == 'Alice'
    assert rows[1]['city'] == 'LA'


def test_XmlReader_auto_row_tag(tmp_path):
    """XmlReader auto-detects the most common tag as row separator."""
    xml_file = tmp_path / 'data.xml'
    xml_file.write_text(_TABULAR_XML)
    r = XmlReader(str(xml_file))
    rows = list(r())
    assert len(rows) == 2
    assert rows[0]['name'] == 'Alice'


def test_XmlReader_filelike(tmp_path):
    """XmlReader accepts a file-like; reads content at construction time."""
    xml_file = tmp_path / 'data.xml'
    xml_file.write_text(_TABULAR_XML)
    with open(str(xml_file)) as f:
        r = XmlReader(f, row_tag='row')
    rows = list(r())
    assert len(rows) == 2
    assert rows[0]['name'] == 'Alice'
    assert rows[1]['name'] == 'Bob'


# ===========================================================================
# getReader — file-like dispatch
# ===========================================================================


def test_getReader_filelike_csv_by_name():
    """getReader dispatches to CsvReader when file-like.name ends in .csv."""
    f = io.StringIO("col_a,col_b\nfoo,bar")
    f.name = 'upload.csv'
    r = getReader(f)
    assert isinstance(r, CsvReader)
    rows = list(r())
    assert rows[0]['col_a'] == 'foo'


def test_getReader_filelike_xls_by_name():
    """getReader dispatches to XlsReader when file-like.name ends in .xls."""
    xls_path = os.path.join(DATA_DIR, 'test.xls')
    with open(xls_path, 'rb') as f:
        data = io.BytesIO(f.read())
    data.name = 'report.xls'
    r = getReader(data)
    assert isinstance(r, XlsReader)
    assert len(list(r())) == 1


def test_getReader_filelike_xlsx_by_name():
    """getReader dispatches to XlsxReader when file-like.name ends in .xlsx."""
    pytest.importorskip('openpyxl')
    xlsx_path = os.path.join(DATA_DIR, 'test.xlsx')
    with open(xlsx_path, 'rb') as f:
        data = io.BytesIO(f.read())
    data.name = 'report.xlsx'
    r = getReader(data)
    assert isinstance(r, XlsxReader)
    assert len(list(r())) == 1


def test_getReader_filelike_explicit_filetype_excel():
    """getReader uses filetype kwarg when file-like has no meaningful name."""
    xls_path = os.path.join(DATA_DIR, 'test.xls')
    with open(xls_path, 'rb') as f:
        data = io.BytesIO(f.read())   # BytesIO has no .name
    r = getReader(data, filetype='excel')
    assert isinstance(r, XlsReader)
    assert len(list(r())) == 1


# ===========================================================================
# XlsWriter round-trip
# ===========================================================================

def test_XlsWriter_round_trip():
    """XlsWriter produces a readable XLS; XlsReader gets the original values back."""
    pytest.importorskip('xlwt')
    pytest.importorskip('xlrd')

    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as f:
        xls_path = f.name
    try:
        w = XlsWriter(
            headers=['name', 'city'],
            columns=['name', 'city'],
            coltypes={'name': 'T', 'city': 'T'},
            filepath=xls_path,
            sheet_base_name='Sheet1',
        )
        w.writeHeaders()
        w.writeRow({'name': 'Alice', 'city': 'NYC'})
        w.writeRow({'name': 'Bob', 'city': 'LA'})
        w.workbookSave()

        r = XlsReader(xls_path)
        assert 'name' in r.headers
        assert 'city' in r.headers
        rows = list(r())
        assert len(rows) == 2
        assert rows[0]['name'] == 'Alice'
        assert rows[0]['city'] == 'NYC'
        assert rows[1]['name'] == 'Bob'
    finally:
        os.unlink(xls_path)


# ===========================================================================
# XlsxWriter round-trip
# ===========================================================================

def test_XlsxWriter_round_trip():
    """XlsxWriter produces a readable XLSX; XlsxReader gets the original values back."""
    pytest.importorskip('openpyxl')

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        xlsx_path = f.name
    try:
        w = XlsxWriter(
            headers=['name', 'city'],
            columns=['name', 'city'],
            coltypes={'name': 'T', 'city': 'T'},
            filepath=xlsx_path,
            sheet_base_name='Sheet1',
        )
        w.writeHeaders()
        w.writeRow({'name': 'Alice', 'city': 'NYC'})
        w.writeRow({'name': 'Bob', 'city': 'LA'})
        w.workbookSave()

        r = XlsxReader(xlsx_path)
        assert 'name' in r.headers
        assert 'city' in r.headers
        rows = list(r())
        assert len(rows) == 2
        assert rows[0]['name'] == 'Alice'
        assert rows[0]['city'] == 'NYC'
        assert rows[1]['name'] == 'Bob'
    finally:
        os.unlink(xlsx_path)


# ===========================================================================
# Legacy functions — smoke tests confirming they still live in flatfiles
# ===========================================================================

def test_legacy_functions_importable_from_flatfiles():
    """readTab, readCSV, readCSV_new, readXLS are accessible from flatfiles."""
    assert callable(readTab)
    assert callable(readCSV)
    assert callable(readCSV_new)
    assert callable(readXLS)

# ===========================================================================
# getCsvDialect / getReader csv_auto — clevercsv-backed dialect detection
# ===========================================================================

def test_getCsvDialect():
    """Test getCsvDialect: detects correct delimiter, quotechar and escapechar for various CSV formats"""
    test_cases = [
        ('test_CsvAuto_Comma.csv', ','),
        ('test_CsvAuto_CommaQuotedDecimalsEUR.csv', ','),
        ('test_CsvAuto_CommaQuotedDecimalsUSA.csv', ','),
        ('test_CsvAuto_SemiColon.csv', ';'),
        ('test_CsvAuto_Tab.csv', '\t'),
        ('test_CsvAuto_Pipe.csv', '|'),
        ('test_CsvAuto_Colon.csv', ':'),
    ]

    for filename, expected_delimiter in test_cases:
        test_file = os.path.join(DATA_DIR, filename)
        dialect = getCsvDialect(test_file, encoding='utf-8')

        assert dialect.delimiter == expected_delimiter
        assert dialect.quotechar == '"'
        assert not dialect.escapechar


def test_getCsvDialect_limited_lines():
    """Test getCsvDialect with detector_max_lines=1 does not detect quotechar

    When reading only the first line, the dialect detector should not
    find any quotechar since the quoted content only appears in later rows.
    """
    test_files = [
        'test_CsvAuto_Comma.csv',
        'test_CsvAuto_CommaQuotedDecimalsEUR.csv',
        'test_CsvAuto_CommaQuotedDecimalsUSA.csv',
        'test_CsvAuto_SemiColon.csv',
        'test_CsvAuto_Tab.csv',
        'test_CsvAuto_Pipe.csv',
        'test_CsvAuto_Colon.csv',
    ]

    for filename in test_files:
        test_file = os.path.join(DATA_DIR, filename)
        dialect = getCsvDialect(test_file, encoding='utf-8', detector_max_lines=1)

        # When reading only first line, quotechar should not be detected
        assert not dialect.quotechar


def test_getReader_CsvAuto():
    """Test getReader with csv_auto filetype detects various delimiters correctly."""
    test_files = [
        'test_CsvAuto_Colon.csv',
        'test_CsvAuto_Comma.csv',
        'test_CsvAuto_CommaQuotedDecimalsEUR.csv',
        'test_CsvAuto_CommaQuotedDecimalsUSA.csv',
        'test_CsvAuto_Pipe.csv',
        'test_CsvAuto_SemiColon.csv',
        'test_CsvAuto_Tab.csv',
    ]

    for filename in test_files:
        test_file = os.path.join(DATA_DIR, filename)

        reader = getReader(test_file, filetype='csv_auto')

        assert reader.ncols == 11
        assert reader.headers[0] == 'Data contabile'
        assert reader.headers[10] == 'Note'
        assert len(list(reader())) == 6


def test_CsvReader_quoted_decimals():
    """Test CsvReader correctly reads quoted decimal values in different formats (EUR vs USA)"""
    test_cases = [
        ('test_CsvAuto_CommaQuotedDecimalsEUR.csv', '-50,00'),  # European format: comma as decimal separator
        ('test_CsvAuto_CommaQuotedDecimalsUSA.csv', '-50.00'),  # US format: period as decimal separator
    ]

    expected_description = 'PAGAMENTO   CARTA DEBITO;\tINTERNAZIONALE: "5375******3179" -03/01/26-13:49 BIANCHI NEGOZIO ABBIGLIAMENTO -ITA'

    for filename, expected_importo in test_cases:
        test_file = os.path.join(DATA_DIR, filename)

        # Detect dialect first, then create CsvReader
        dialect = getCsvDialect(test_file, encoding='utf-8')
        reader = CsvReader(test_file, dialect=dialect, encoding='utf-8')

        assert reader.ncols == 11
        assert reader.headers[0] == 'Data contabile'
        assert reader.headers[10] == 'Note'

        rows = list(reader())
        assert len(rows) == 6

        last_row = rows[5]
        assert last_row[2] == expected_importo
        assert last_row[9] == expected_description


def test_CsvReader_start_at_line():
    """Test CsvReader start_at_line parameter skips header lines correctly.

    Uses test_CsvAuto_Colon_skipLines.csv which has 12 lines of metadata before the actual CSV data.
    With start_at_line=12, results should be identical to test_CsvAuto_Colon.csv without the parameter.
    """
    # Reference file (no skip)
    reference_file = os.path.join(DATA_DIR, 'test_CsvAuto_Colon.csv')
    dialect = getCsvDialect(reference_file, encoding='utf-8')
    reference_reader = CsvReader(reference_file,
                                 dialect=dialect, encoding='utf-8')
    reference_rows = list(reference_reader())

    # File with metadata to skip
    START_LINE = 12
    skip_file = os.path.join(DATA_DIR, 'test_CsvAuto_Colon_skipLines.csv')
    dialect = getCsvDialect(skip_file, encoding='utf-8',
                            start_at_line=START_LINE)
    skip_reader = CsvReader(skip_file,
                            dialect=dialect, encoding='utf-8',
                            start_at_line=START_LINE)
    skip_rows = list(skip_reader())

    assert reference_reader.headers == skip_reader.headers
    assert len(reference_rows) == len(skip_rows)

    for ref_row, skip_row in zip(reference_rows, skip_rows):
        for j in range(len(ref_row)):
            assert ref_row[j] == skip_row[j]


def test_CsvReader_auto_dialect():
    """Test CsvReader with automatic dialect detection via getCsvDialect.

    Verifies that CsvReader correctly reads CSV files with various delimiters
    (comma, semicolon, tab, pipe, colon) when dialect is detected via getCsvDialect.
    """
    # Test files with different delimiters and decimal formats
    test_files = [
        ('test_CsvAuto_Colon.csv', '-50,00'),
        ('test_CsvAuto_Comma.csv', '-50.00'),
        ('test_CsvAuto_CommaQuotedDecimalsEUR.csv', '-50,00'),
        ('test_CsvAuto_CommaQuotedDecimalsUSA.csv', '-50.00'),
        ('test_CsvAuto_Pipe.csv', '-50,00'),
        ('test_CsvAuto_SemiColon.csv', '-50,00'),
        ('test_CsvAuto_Tab.csv', '-50,00'),
    ]

    expected_description = 'PAGAMENTO   CARTA DEBITO;\tINTERNAZIONALE: "5375******3179" -03/01/26-13:49 BIANCHI NEGOZIO ABBIGLIAMENTO -ITA'

    for filename, expected_importo in test_files:
        test_file = os.path.join(DATA_DIR, filename)

        # Detect dialect first, then create CsvReader
        dialect = getCsvDialect(test_file, encoding='utf-8')
        reader = CsvReader(test_file, dialect=dialect, encoding='utf-8')

        assert reader.ncols == 11
        assert reader.headers[0] == 'Data contabile'
        assert reader.headers[10] == 'Note'

        rows = list(reader())
        assert len(rows) == 6

        last_row = rows[5]
        assert last_row[2] == expected_importo
        assert last_row[9] == expected_description


### ported from gnrlist_test
def test_getReader():

    with tempfile.TemporaryDirectory() as tmpdir:
        filename = os.path.join(tmpdir, 'test.csv')
        with open(filename, "w") as wfp:
            wfp.write("one,two,three\nfour,five,six")
        a = getReader(filename)
        assert isinstance(a, CsvReader)
        a.filecsv.close()

        filename = os.path.join(tmpdir, 'test.tab')
        with open(filename, "w") as wfp:
            wfp.write("one\ttwo\tthree\nfour\tfive\tsix")
        a = getReader(filename)
        assert isinstance(a, CsvReader)
        a.filecsv.close()
        
        filename = os.path.join(tmpdir, 'test.csv')
        with open(filename, "w") as wfp:
            wfp.write("one\ttwo\tthree\nfour\tfive\tsix")
        a = getReader(filename, filetype="csv_auto")
        assert isinstance(a, CsvReader)
        a.filecsv.close()

        # Will fail with an emtpy file
        filename = os.path.join(tmpdir, 'test.xls')
        with open(filename, "w") as wfp:
            pass
        with pytest.raises(Exception):
            a = getReader(filename, filetype="excel")
            a.filecsv.close()
        filename = os.path.join(tmpdir, 'test.xlsx')
        with open(filename, "w") as wfp:
            pass
        with pytest.raises(Exception):
            a = getReader(filename, filetype="excel")
            a.filecsv.close()
            
    test_dir = os.path.dirname(__file__)
    
    filename = os.path.join(test_dir, "data", "test.xls")
    a = getReader(filename)
    assert isinstance(a, XlsReader)

    filename = os.path.join(test_dir, "data","test.xlsx")
    a = getReader(filename)
    assert isinstance(a, XlsxReader)

    # FIXME: this fails all the time.
    with pytest.raises(Exception):
        filename = os.path.join(test_dir, "data", "testbag.xml")
        a = getReader(filename)

def test_CsvReader():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", "test.csv")
    a = CsvReader(test_file)
    # FIXME: odd interface using __call__
    r = [x for x in a()]
    assert len(r) == 1
    assert isinstance(r[0], GnrNamedList)
    assert 'a' in r[0].keys()
    a = CsvReader(test_file, detect_encoding=True)


def test_XlsReader():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.xls')
    r = XlsReader(test_file)
    assert r.sheet.name == "Sheet1"
    assert 'a' in r.headers
    assert 0 in r.colindex
    assert r.colindex[0] is True
    assert 'a' in r.index
    assert r.ncols == 3
    assert r.nrows == 1
    d = [x for x in r()]
    assert len(d) == 1
    assert isinstance(d[0], GnrNamedList)
    assert 'a' in d[0].keys()

def test_XlsxReader():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.xlsx')
    r = XlsxReader(test_file)
    assert r.sheet.title == "Sheet1"
    assert 'a' in r.headers
    assert 0 in r.colindex
    assert r.colindex[0] is True
    assert 'a' in r.index
    assert r.ncols == 3

    # FIXME: this doesn't work in the implementation
    #assert r.nrows == 1

    d = [x for x in r()]
    assert len(d) == 1
    assert isinstance(d[0], GnrNamedList)
    assert 'a' in d[0].keys()

def test_readXLS():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.xls')
    r = readXLS(test_file)
    d = [x for x in r]
    assert len(d) == 1
    assert isinstance(d[0], GnrNamedList)
    assert 'a' in d[0].keys()

    with open(test_file, "rb") as fp:
        r = readXLS(fp)
        d = [x for x in r]
        assert len(d) == 1
        assert isinstance(d[0], GnrNamedList)
        assert 'a' in d[0].keys()

def test_readCSV():
    # FIXME: apparently, readXLS and readCSV exposes
    # a different interface to access record, please
    # check the last assert here with the last of readXLS test
    # while readCSV_new works correctly. Maybe the _new should
    # be the implementation..
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.csv')
    r = readCSV(test_file)
    d = [x for x in r]
    assert len(d) == 2
    assert isinstance(d[0], GnrNamedList)
    assert 'a' in d[0].keys()[0]

    with open(test_file, "r") as fp:
        r = readCSV(fp)
        d = [x for x in r]
        assert len(d) == 2
        assert isinstance(d[0], GnrNamedList)
        assert 'a' in d[0].keys()[0]

def test_readCSV_new():
    test_dir = os.path.dirname(__file__)
    test_file = os.path.join(test_dir, "data", 'test.csv')
    r = readCSV_new(test_file)
    d = [x for x in r]
    assert len(d) == 1
    assert isinstance(d[0], GnrNamedList)
    assert 'a' in d[0].keys()

    with open(test_file, "r") as fp:
        r = readCSV_new(fp)
        d = [x for x in r]
        assert len(d) == 1
        assert isinstance(d[0], GnrNamedList)
        assert 'a' in d[0].keys()

def test_CsvReader_duplicate_columns():
    """Test handling of duplicate column names in CSV files"""

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        csv_file = f.name
        writer = csv.writer(f)
        # Headers with duplicate 'name' column
        writer.writerow(['id', 'name', 'surname', 'name', 'email'])
        writer.writerow(['1', 'Mario', 'Rossi', 'Giuseppe', 'mario@test.com'])
        writer.writerow(['2', 'Laura', 'Bianchi', 'Anna', 'laura@test.com'])

    try:
        reader = CsvReader(csv_file)

        # Check that duplicate column has been renamed
        assert 'name' in reader.headers
        assert 'name[3]' in reader.headers
        assert reader.headers == ['id', 'name', 'surname', 'name[3]', 'email']

        # Check index mapping
        assert reader.index['name'] == 1
        assert reader.index['name[3]'] == 3

        # Read rows
        rows = [row for row in reader()]
        assert len(rows) == 2

        # Test first row
        row = rows[0]
        assert row[0] == '1'
        assert row[1] == 'Mario'
        assert row[2] == 'Rossi'
        assert row[3] == 'Giuseppe'
        assert row[4] == 'mario@test.com'

        # Access by name
        assert row['id'] == '1'
        assert row['name'] == 'Mario'
        assert row['surname'] == 'Rossi'
        assert row['name[3]'] == 'Giuseppe'
        assert row['email'] == 'mario@test.com'

    finally:
        os.unlink(csv_file)


def test_XlsxReader_duplicate_columns():
    """Test handling of duplicate column names in XLSX files"""
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        xlsx_file = f.name

    try:
        # Create XLSX with duplicate columns
        wb = Workbook()
        ws = wb.active
        ws.append(['id', 'name', 'surname', 'name', 'email'])
        ws.append(['1', 'Mario', 'Rossi', 'Giuseppe', 'mario@test.com'])
        ws.append(['2', 'Laura', 'Bianchi', 'Anna', 'laura@test.com'])
        wb.save(xlsx_file)

        reader = XlsxReader(xlsx_file)

        # Check that duplicate column has been renamed
        assert 'name' in reader.headers
        assert 'name[3]' in reader.headers
        assert reader.headers == ['id', 'name', 'surname', 'name[3]', 'email']

        # Check index mapping
        assert reader.index['name'] == 1
        assert reader.index['name[3]'] == 3

        # Read rows
        rows = [row for row in reader()]
        assert len(rows) == 2

        # Test first row
        row = rows[0]
        assert row[0] == '1'
        assert row[1] == 'Mario'
        assert row[2] == 'Rossi'
        assert row[3] == 'Giuseppe'
        assert row[4] == 'mario@test.com'

        # Access by name
        assert row['id'] == '1'
        assert row['name'] == 'Mario'
        assert row['surname'] == 'Rossi'
        assert row['name[3]'] == 'Giuseppe'
        assert row['email'] == 'mario@test.com'

    finally:
        os.unlink(xlsx_file)


def test_XlsReader_duplicate_columns():
    """Test handling of duplicate column names in XLS files"""
    try:
        import xlwt
    except ImportError:
        pytest.skip("xlwt not available")

    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as f:
        xls_file = f.name

    try:
        # Create XLS with duplicate columns
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Sheet1')

        # Headers
        headers = ['id', 'name', 'surname', 'name', 'email']
        for col, header in enumerate(headers):
            ws.write(0, col, header)

        # Data rows
        data = [
            ['1', 'Mario', 'Rossi', 'Giuseppe', 'mario@test.com'],
            ['2', 'Laura', 'Bianchi', 'Anna', 'laura@test.com']
        ]
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data):
                ws.write(row_idx, col_idx, value)

        wb.save(xls_file)

        reader = XlsReader(xls_file)

        # Check that duplicate column has been renamed
        assert 'name' in reader.headers
        assert 'name[3]' in reader.headers
        assert reader.headers == ['id', 'name', 'surname', 'name[3]', 'email']

        # Check index mapping
        assert reader.index['name'] == 1
        assert reader.index['name[3]'] == 3

        # Read rows
        rows = [row for row in reader()]
        assert len(rows) == 2

        # Test first row
        row = rows[0]
        assert row[0] == '1'
        assert row[1] == 'Mario'
        assert row[2] == 'Rossi'
        assert row[3] == 'Giuseppe'
        assert row[4] == 'mario@test.com'

        # Access by name
        assert row['id'] == '1'
        assert row['name'] == 'Mario'
        assert row['surname'] == 'Rossi'
        assert row['name[3]'] == 'Giuseppe'
        assert row['email'] == 'mario@test.com'

    finally:
        os.unlink(xls_file)
def test_multiple_duplicate_columns():
    """Test handling of 3+ duplicate columns"""

    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        csv_file = f.name
        writer = csv.writer(f)
        # Three 'name' columns
        writer.writerow(['id', 'name', 'name', 'name', 'email'])
        writer.writerow(['1', 'First', 'Middle', 'Last', 'test@test.com'])

    try:
        reader = CsvReader(csv_file)

        # Check all duplicate columns are renamed
        assert reader.headers == ['id', 'name', 'name[2]', 'name[3]', 'email']
        assert reader.index['name'] == 1
        assert reader.index['name[2]'] == 2
        assert reader.index['name[3]'] == 3

        rows = list(reader())
        row = rows[0]

        # All values accessible
        assert row['name'] == 'First'
        assert row['name[2]'] == 'Middle'
        assert row['name[3]'] == 'Last'

    finally:
        os.unlink(csv_file)
def test_slugify_consistency_across_readers():
    """Test that CSV, XLS, and XLSX readers use consistent slugification"""

    # Headers with spaces and special characters
    headers = ['Transaction ID', 'User Name', 'Email-Address', 'Created At']
    expected_keys = ['transaction_id', 'user_name', 'email_address', 'created_at']

    # Test CSV via getReader
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as f:
        csv_file = f.name
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerow(['123', 'Alice', 'alice@test.com', '2025-10-28'])

    try:
        csv_reader = getReader(csv_file)
        csv_keys = list(csv_reader.index.keys())
        assert csv_keys == expected_keys, f"CSV keys: {csv_keys}"

        # Verify access works with underscores
        for row in csv_reader():
            assert row['transaction_id'] == '123'
            assert row['user_name'] == 'Alice'
            break
    finally:
        os.unlink(csv_file)

    # Test XLSX
    try:
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not available")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        xlsx_file = f.name

    try:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(['123', 'Alice', 'alice@test.com', '2025-10-28'])
        wb.save(xlsx_file)

        xlsx_reader = XlsxReader(xlsx_file)
        xlsx_keys = list(xlsx_reader.index.keys())
        assert xlsx_keys == expected_keys, f"XLSX keys: {xlsx_keys}"

        # Verify access works with underscores
        for row in xlsx_reader():
            assert row['transaction_id'] == '123'
            assert row['user_name'] == 'Alice'
            break
    finally:
        os.unlink(xlsx_file)

    # Test XLS
    try:
        import xlwt
    except ImportError:
        pytest.skip("xlwt not available")

    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as f:
        xls_file = f.name

    try:
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Sheet1')
        for col, header in enumerate(headers):
            ws.write(0, col, header)
        for col, value in enumerate(['123', 'Alice', 'alice@test.com', '2025-10-28']):
            ws.write(1, col, value)
        wb.save(xls_file)

        xls_reader = XlsReader(xls_file)
        xls_keys = list(xls_reader.index.keys())
        assert xls_keys == expected_keys, f"XLS keys: {xls_keys}"

        # Verify access works with underscores
        for row in xls_reader():
            assert row['transaction_id'] == '123'
            assert row['user_name'] == 'Alice'
            break
    finally:
        os.unlink(xls_file)

def test_readTab():
    """Test tab-delimited file reading"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.tab', delete=False) as f:
        tab_file = f.name
        f.write("name\tage\tcity\n")
        f.write("Alice\t30\tNYC\n")
        f.write("Bob\t25\tLA\n")

    try:
        rows = list(readTab(tab_file))
        assert len(rows) == 2
        assert isinstance(rows[0], GnrNamedList)
        assert rows[0]['name'] == 'Alice'
        assert rows[0]['age'] == '30'
        assert rows[0]['city'] == 'NYC'
        assert rows[1]['name'] == 'Bob'

    finally:
        os.unlink(tab_file)
